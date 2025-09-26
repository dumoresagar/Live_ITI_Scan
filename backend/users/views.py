from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from .forms import LoginForm, UploadFileForm
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from documents.models import Files,ReceivedDocuments
from django.db.models import Count, Q, Sum, Case, When, IntegerField, Max
from django.utils import timezone
from datetime import timedelta
from .models import UserActivity, Zone, District, Office, Agency
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl
from django.utils.timezone import now, make_aware, get_current_timezone,timedelta
from django.core.exceptions import ObjectDoesNotExist
from pathlib import Path
import json
from PIL import Image
from rest_framework.response import Response
from rest_framework.generics import GenericAPIView
from .serializer import LoginSerializer
from io import BytesIO
from rest_framework.authtoken.models import Token
from rest_framework.permissions import AllowAny
import io
import xlsxwriter
import requests
from openpyxl.utils import get_column_letter
User = get_user_model()

from datetime import datetime

import pandas as pd
from .forms import ExcelUploadForm,DistrictUserUploadForm,DigrUserUploadForm

def upload_excel_srocode(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            try:
                # Read Excel File
                df = pd.read_excel(excel_file, dtype=str)
                df.columns = df.columns.str.strip()  # Clean column names
                
                for _, row in df.iterrows():
                    zone_name = row["zone"].strip()
                    district_name = row["district"].strip()
                    sro_name = row["sro_name"].strip().upper()  # Match case-insensitively
                    sro_id = row["srocode"].strip() if pd.notna(row["srocode"]) else None
                    
                    # Format Office Code
                    if sro_id and sro_id.isdigit():
                        sro_id = sro_id.zfill(3) if len(sro_id) < 4 else sro_id
                    # Get or Create Zone
                    zone, _ = Zone.objects.get_or_create(zone_name=zone_name)
                    # Get or Create District
                    district, _ = District.objects.get_or_create(
                        zone=zone, district_name=district_name
                    )
                    # Find Office (Case-Insensitive)
                    office = Office.objects.filter(office_name__iexact=sro_name, district=district).first()

                    if office:
                        # Update Missing Office Code
                        if not office.office_code and sro_id:
                            office.office_code = sro_id
                            office.save()
                    else:
                        # Create New Office
                        Office.objects.create(
                            office_name=sro_name,
                            office_code=sro_id,
                            district=district
                        )

                messages.success(request, "Excel file uploaded and processed successfully!")
                return redirect('upload_excel')
            
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
                return redirect('upload_excel')

    else:
        form = UploadFileForm()

    return render(request, 'users/upload_excel.html', {'form': form})

def count_tiff_pages(file_path):
    try:
        with Image.open(file_path) as img:
            return img.n_frames  # n_frames gives the total number of frames/pages in the TIFF
    except Exception as e:
        # print(f"Error counting pages for {file_path}: {e}")
        return 0

def upload_excel_sro(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Load the Excel file
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
        except Exception as e:
            messages.error(request, f"Error reading the Excel file: {e}")
            return redirect('upload_excel')

        # Iterate over the rows in the Excel sheet (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            office_name, address, telephone, email = row
            
            try:
                office = Office.objects.get(office_name=office_name)
            except ObjectDoesNotExist:
                messages.warning(request, f"Office {office_name} not found, skipping user creation.")
                continue

            # Check if user already exists
            if User.objects.filter(email=email).exists():
                messages.warning(request, f"User with email {email} already exists. Skipping.")
                continue

            # Create the user
            user = User.objects.create_user(
                username=email,
                email=email,
                password='Sro@123',  # Default password
                first_name=office_name,  # Default first name (you can modify as needed)
                last_name='',
                is_department=True,
                is_active=True,
                office=office,
                contact=telephone,
                avatar=None,  # Default avatar (you can modify as needed)
            )

            messages.success(request, f"User {email} created successfully!")

        return HttpResponse("Data uploaded successfully!", content_type="text/plain")

    form = ExcelUploadForm()
    return render(request, 'users/upload_sro.html', {'form': form})

def upload_district_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Load the Excel file
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
        except Exception as e:
            messages.error(request, f"Error reading the Excel file: {e}")
            return redirect('upload_district_excel')

        # Iterate over the rows in the Excel sheet (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            district_name, address, telephone, email_ids = row
            
            # Split the email IDs by space
            email_list = email_ids.split()
            
            # Try to get the district object
            try:
                district = District.objects.get(district_name=district_name.upper())
            except ObjectDoesNotExist:
                messages.warning(request, f"District {district_name} not found, skipping user creation.")
                continue

            # Loop through the emails and create users
            for index, email in enumerate(email_list):
                # Check if the user already exists
                if User.objects.filter(email=email).exists():
                    messages.warning(request, f"User with email {email} already exists. Skipping.")
                    continue

                # Create the user
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password='Dro@123',  # Default password
                    first_name=f"{district_name}{index + 1}",  # First name as DistrictName1, DistrictName2, etc.
                    last_name='',
                    is_district_rgtr=True,  # Flag for district registrar
                    is_active=True,
                    district=district,
                    contact=telephone,
                    avatar=None,  # Default avatar if needed
                )

                messages.success(request, f"User {email} created successfully!")

        return redirect('upload_district_excel')

    form = DistrictUserUploadForm()
    return render(request, 'users/upload_district_excel.html', {'form': form})

def upload_digr_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Load the Excel file
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
        except Exception as e:
            messages.error(request, f"Error reading the Excel file: {e}")
            return redirect('upload_digr_excel')

        # Iterate over the rows in the Excel sheet (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            zone_name, address, telephone, email = row
            
            # Try to get the zone object
            try:
                zone = Zone.objects.get(zone_name=zone_name.upper())
            except ObjectDoesNotExist:
                messages.warning(request, f"Zone {zone_name} not found, skipping user creation.")
                continue

            # Check if the user already exists
            if User.objects.filter(email=email).exists():
                messages.warning(request, f"User with email {email} already exists. Skipping.")
                continue

            # Create the user
            user = User.objects.create_user(
                username=email,
                email=email,
                password='Dig@123',  # Default password
                first_name=zone_name,  # Default first name (you can modify as needed)
                last_name='',
                is_digr=True,  # Flag for DIGR
                is_active=True,
                zone=zone,
                contact=telephone,
                avatar=None,  # Default avatar if needed
            )

            messages.success(request, f"User {email} created successfully!")

        return HttpResponse("Data uploaded successfully!", content_type="text/plain")

    form = DigrUserUploadForm()
    return render(request, 'users/upload_digr_excel.html', {'form': form})


def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        # Load the Excel file
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file, engine='openpyxl')

        # Process each row in the Excel file
        for index, row in df.iterrows():
            # Get or create the Zone
            zone_name = row['ZONE']
            zone, created = Zone.objects.get_or_create(zone_name=zone_name)

            # Get or create the District
            district_name = row['DISTRICT']
            district, created = District.objects.get_or_create(zone=zone, district_name=district_name)

            # Get or create the Office
            office_name = row['OFFICE NAME']
            office, created = Office.objects.get_or_create(district=district, office_name=office_name)

        return HttpResponse("Data uploaded successfully!", content_type="text/plain")  # Or another template to show success message
    return render(request, 'users/upload.html')  # A template to upload the Excel file



class LoginView(GenericAPIView):
    serializer_class = LoginSerializer
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.validated_data['user']   

            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])

            if user.agency:
                office_code = user.office.office_code
            else:
                office_code = None

            # Create or get token
            token, created = Token.objects.get_or_create(user=user)

            UserActivity.objects.create(
                user=user,
                action='login',
                ip_address=request.ip_address,
                device_type=request.device_type,
                browser=request.browser,
                os=request.os,
                details="User logged in successfully"
            )

            return Response({
                "message": "Login Successful",
                "token": token.key,
                "user": {
                    "username": user.username,
                    "id": user.id,
                    "full_name": user.get_full_name(),
                    "office_code": office_code
                }
            })
        else:
            return Response({"error": "Invalid username or Password"}, status=400)



def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_igr:
                return redirect('igr_dash')
        else:
            return redirect('dashboard')
        
    form = LoginForm(request.POST or None) 
    
    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data['username']
        password = form.cleaned_data['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user) 
            messages.success(request, 'Login successful!')
            UserActivity.objects.create(
                                        user=request.user,
                                        action='login',
                                        ip_address=request.ip_address,
                                        device_type=request.device_type,
                                        browser=request.browser,
                                        os=request.os,
                                        details="User logged in successfully"
                                    )
            if user.is_igr:
                return redirect('igr_dash')
            elif user.is_agency_scanning_employee:
                messages.error(request, 'You are not allowed to login Please contact admin.')
            else:
                return redirect('dashboard')
            
            
        else:
            messages.error(request, 'Invalid username or password or Your account is not active.')

    return render(request, 'users/login.html', {'form': form})


def create_agency_qc_users_excel(request):
    try:
        admin_user = User.objects.get(username='dalvkot_admin@mail.com')
        agency = admin_user.agency
    except User.DoesNotExist:
        return HttpResponse("Agency admin with username 'dalvkot_admin@mail.com' not found.", status=404)
    except AttributeError:
        return HttpResponse("Admin user does not have an agency assigned.", status=400)

    # Store user info for Excel
    user_data = []

    default_password = "dalvkot@123"

    for office in Office.objects.all():
        for i in range(1, 3):
            username = f"{office.office_name.lower().replace(' ', '')}{i}@dalvkot.com"

            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(
                    username=username,
                    email=username,
                    password=default_password,
                    first_name=office.office_name,
                    last_name=f"user {i}",
                    office=office,
                    agency=agency,
                    is_agency_qc_employee=True,
                    is_agency=True,
                )
                user_data.append([office.office_name, username, default_password])
            else:
                # You may include already existing users if needed
                pass

    # Generate Excel file in memory
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Agency QC Users")

    # Write headers
    headers = ['Office Name', 'Username', 'Password']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write user rows
    for row_num, row_data in enumerate(user_data, start=1):
        for col_num, value in enumerate(row_data):
            worksheet.write(row_num, col_num, value)

    workbook.close()
    output.seek(0)

    # Send as response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=agency_qc_users.xlsx'

    return response

@login_required
def logout_view(request):
    # Capture the current user before logout
    user = request.user

    # Log the logout activity before the user is logged out
    UserActivity.objects.create(
        user=user,
        action='logout',  # Corrected action from 'login' to 'logout'
        ip_address=request.ip_address,
        device_type=request.device_type,
        browser=request.browser,
        os=request.os,
        details="User logged out successfully"
    )

    logout(request)

    if 'token' in request.session:
        del request.session['token']

    messages.info(request, 'You have been logged out successfully.')
    return redirect('login')


from django.contrib.admin.models import LogEntry

@login_required
def dashboard(request):
    if request.user.is_igr:
        return redirect('igr_dash')
    filter_option = request.GET.get('filter', 'today')

    # Get current time in the user's timezone
    now = timezone.localtime(timezone.now())  # This is already timezone-aware

    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'month':
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'year':
        start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start_date = None

    # Check if start_date is naive, and make it aware if it is
    if start_date and timezone.is_naive(start_date):
        start_date = timezone.make_aware(start_date, timezone.get_current_timezone())

    # Filter the activities based on the time range
    if start_date:
        activities = UserActivity.objects.filter(user=request.user, timestamp__gte=start_date)
    else:
        activities = UserActivity.objects.filter(user=request.user)

            
    today = timezone.now()
    seven_days_ago = today - timedelta(days=7)
    
    admin_qc_files = Files.objects.filter(processed=True,send_to_sro = False,send_to_qc= True,admin_approved = None, dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    if request.user.is_agency_qc_employee:
        admin_qc_files = Files.objects.filter(uploaded_by__office = request.user.office,uploaded_by__agency = request.user.agency ,processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at')
    if request.user.is_agency_admin:
        admin_qc_files = Files.objects.filter(uploaded_by__agency = request.user.agency ,processed=True,send_to_qc = True).order_by('-uploaded_at')

    p_files = Files.objects.filter(office=request.user.office,processed=True,send_to_sro = True, dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    u_files = Files.objects.filter(processed=False,send_to_sro = True,dept_approved = None)
    d_a_files = Files.objects.filter(processed=True,send_to_sro = True,dept_approved = True)
    d_r_files = (
                    Files.objects.filter(
                        Q(processed=True,send_to_sro = True, dept_approved=False) |
                        Q(processed=True,send_to_sro = True, district_rgtr_approved=False) |
                        Q(processed=True,send_to_sro = True, digr_approved=False)
                    )
                    .order_by('uploaded_at')
                )

    sro_r_files = Files.objects.filter(office = request.user.office,processed=True, dept_approved=False)
    dist_a_files =  Files.objects.filter(office__district = request.user.district ,processed = True, dept_approved = True,district_rgtr_approved = None)
    dist_approved_files =  Files.objects.filter(office__district = request.user.district ,processed = True, dept_approved = True,district_rgtr_approved = True)
    dist_rejected_files =  Files.objects.filter(office__district = request.user.district ,processed = True, dept_approved = True,district_rgtr_approved = False)

    dist_r_files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = False)
    dr_a_files = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None)
    digr_r_files = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=False)
    self_digr_r_files = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=False)
    
    sro_self_approved = Files.objects.filter(office = request.user.office,processed=True, dept_approved=True)
    sro_self_pending = Files.objects.filter(office = request.user.office,processed=True,send_to_sro = True, dept_approved=None)
    sro_self_total = Files.objects.filter(office = request.user.office,processed=True,send_to_sro = True)
    sro_self_rejected = Files.objects.filter(office = request.user.office,processed=True,send_to_sro = True, dept_approved=False)
    


    all_approved_files =  Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=True)
    
    # agency 
    
    approved_files_agency = Files.objects.filter(uploaded_by__agency=request.user.agency,
                                                      processed=True,send_to_sro = False,send_to_qc= True,
                                                      admin_approved = True,admin_approved_by = request.user,
                                                      dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    
    rejected_files_agency = Files.objects.filter(uploaded_by__agency=request.user.agency,
                                                      processed=True,send_to_sro = False,send_to_qc= True,
                                                      admin_approved = False,admin_approved_by = request.user,
                                                      dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    
    
    
    self_approved_files_agency = Files.objects.filter(uploaded_by__office = request.user.office,uploaded_by__agency=request.user.agency,
                                                      processed=True,send_to_sro = False,send_to_qc= True,
                                                      admin_approved = True,admin_approved_by = request.user,
                                                      dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    
    self_rejected_files_agency = Files.objects.filter(uploaded_by__office = request.user.office,uploaded_by__agency=request.user.agency,
                                                      processed=True,send_to_sro = False,send_to_qc= True,
                                                      admin_approved = False,admin_approved_by = request.user,
                                                      dept_approved=None, district_rgtr_approved=None,digr_approved=None)
    
    dig_self_approved_files =  Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=True)
    
    
    search_query = request.GET.get('q', '')
    selected_document_types = request.GET.getlist('document_type')  # Get multiple selected types
    from_date = request.GET.get('from_date', seven_days_ago)
    to_date = request.GET.get('to_date', today)

    # Base queryset for the last 7 days if no date filters are provided
    sro_data = Files.objects.filter(dept_approved_at__gte=seven_days_ago)

    # Apply date range filters if provided
    if from_date:
        sro_data = sro_data.filter(dept_approved_at__date__gte=from_date)
    if to_date:
        sro_data = sro_data.filter(dept_approved_at__date__lte=to_date)

    # Apply other filters
    if search_query:
        sro_data = sro_data.filter(filename__icontains=search_query)
    if selected_document_types:
        query = Q()
        for doc_type in selected_document_types:
            query |= Q(filename__startswith=doc_type)
        sro_data = sro_data.filter(query)
    

    # Aggregations
    total_files = sro_data.count()
    approved_files = sro_data.filter(dept_approved=True).count()
    rejected_files = sro_data.filter(dept_approved=False).count()

    # Distinct values for filters
    document_types = set(
        f.split('_')[0] for f in Files.objects.values_list('filename', flat=True)
    )
    office_codes = set(
        f.split('_')[1]
        for f in Files.objects.values_list('filename', flat=True)
        if len(f.split('_')) > 1
    )
    years = set(
        f.split('_')[-1]
        for f in Files.objects.values_list('filename', flat=True)
        if f.split('_')[-1].isdigit()
    )
    
    try:
        if isinstance(from_date, str) and from_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
        else:
            from_date = seven_days_ago

        if isinstance(to_date, str) and to_date:
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        else:
            to_date = today

        # Convert dates to timezone-aware if necessary
        if timezone.is_naive(from_date):
            from_date = make_aware(from_date, get_current_timezone())
        if timezone.is_naive(to_date):
            to_date = make_aware(to_date, get_current_timezone())

    except ValueError as e:
        # Handle invalid date formats
        from_date = seven_days_ago
        to_date = today


    
    
    document_type_filter = Q()
    for doc_type in selected_document_types:
        document_type_filter |= Q(filename__startswith=doc_type)

    # Query for SRO data
    sro_dash_data = (
        Files.objects
        .filter(
            office = request.user.office,
            uploaded_at__date__gte=from_date,
            uploaded_at__date__lte=to_date
        )
        .filter(document_type_filter)  # Apply the document type filter
        .values('uploaded_at__date')
        .annotate(
            qc_count=Count('id', filter=Q(send_to_sro=True, dept_approved=None)),
            sro_approved_count=Count('id', filter=Q(dept_approved=True))
        )
        .order_by('-uploaded_at__date')
    )

    # Query for DR data
    dr_dash_data = (
        Files.objects
        .filter(
            dept_approved_by__office__district = request.user.district,
            dept_approved_at__date__gte=from_date,
            dept_approved_at__date__lte=to_date
        )
        .filter(document_type_filter)  # Apply the document type filter
        .values('uploaded_at__date')
        .annotate(
            qc_count=Count('id', filter=Q(district_rgtr_approved=None, dept_approved=True)),
            dr_approved_count=Count('id', filter=Q(district_rgtr_approved=True))
        )
        .order_by('-uploaded_at__date')
    )

    # Query for DIGR data
    digr_dash_data = (
        Files.objects
        .filter(
            district_rgtr_approved_by__district__zone = request.user.zone,
            dept_approved_at__date__gte=from_date,
            dept_approved_at__date__lte=to_date
        )
        .filter(document_type_filter)  # Apply the document type filter
        .values('uploaded_at__date')
        .annotate(
            qc_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),
            digr_approved_count=Count('id', filter=Q(digr_approved=True))
        )
        .order_by('-uploaded_at__date')
    )
    
    
        
    offices_without_code = Office.objects.filter(office_code__isnull=True)

    # Print office names
    # for office in offices_without_code:
    #     print('user__uploaded_files',office.office_name)
    
    agency_data = Agency.objects.all().annotate(
        total_uploaded=Count('user__uploaded_files'),
        sent_to_qc=Count(
            'user__uploaded_files',
            filter=Q(user__uploaded_files__admin_approved=True,
                     user__uploaded_files__dept_approved__isnull=True,
                     user__uploaded_files__district_rgtr_approved__isnull=True,
                     user__uploaded_files__digr_approved__isnull=True)
        ),
        pending=Count(
            'user__uploaded_files',
            filter=Q(user__uploaded_files__district_rgtr_approved__isnull=True) &
                   Q(user__uploaded_files__digr_approved__isnull=True) &
                   Q(user__uploaded_files__dept_approved__isnull=True) &
                   Q(user__uploaded_files__admin_approved__isnull=True)
        ),
        approved_by_sro=Count(
            'user__uploaded_files',
            filter=Q(user__uploaded_files__dept_approved=True,
                     user__uploaded_files__district_rgtr_approved__isnull=True)
        ),
        approved_by_dr=Count(
            'user__uploaded_files',
            filter=Q(user__uploaded_files__district_rgtr_approved=True,
                     user__uploaded_files__digr_approved__isnull=True)
        ),
        approved_by_digr=Count(
            'user__uploaded_files',
            filter=Q(user__uploaded_files__digr_approved=True)
        ),
        rejected=Count(
            'user__uploaded_files',
            filter=(
                Q(user__uploaded_files__district_rgtr_approved=False) |
                Q(user__uploaded_files__digr_approved=False) |
                Q(user__uploaded_files__dept_approved=False) |
                Q(user__uploaded_files__admin_approved=False)
            )
        )
    )
    
    context = {
        'p_files': p_files,
        'u_files': u_files,
        'd_a_files': d_a_files,
        'd_r_files': d_r_files,
        'dist_a_files':dist_a_files,
        'dist_r_files':dist_r_files,
        'dr_a_files':dr_a_files,
        'digr_r_files':digr_r_files,
        'all_approved_files':all_approved_files,
        'sro_dash_data': sro_dash_data,
        'dr_dash_data': dr_dash_data,
        'digr_dash_data': digr_dash_data,
        'seven_days_ago': seven_days_ago,
        'activities': activities,
        'filter_option': filter_option,
        'total_files': total_files,
        'approved_files': approved_files,
        'rejected_files': rejected_files,
        'document_types': sorted(document_types),
        'office_codes': sorted(office_codes),
        'years': sorted(years),
        'selected_document_types': selected_document_types,
        'sro_r_files':sro_r_files,
        'admin_qc_files':admin_qc_files,
        'agency_data': agency_data,
        'sro_self_approved':sro_self_approved,
        'sro_self_rejected':sro_self_rejected,
        'sro_self_pending':sro_self_pending,
        'sro_self_total':sro_self_total,
        "dist_approved_files": dist_approved_files,
        'dist_rejected_files':dist_rejected_files,
        'self_digr_r_files':self_digr_r_files,
        'dig_self_approved_files':dig_self_approved_files,
        'self_approved_files_agency':self_approved_files_agency,
        'self_rejected_files_agency':self_rejected_files_agency,
        'approved_files_agency':approved_files_agency,
        'rejected_files_agency':rejected_files_agency,
        
    }
    return render(request, 'users/dashboard.html',context)




@login_required
def igr_dashboard(request):
    
    all_submitted = Files.objects.filter(send_to_sro = True)
    all_pending = Files.objects.filter(
                    Q(processed=True, dept_approved=True) |
                    Q(processed=True, district_rgtr_approved=True) |
                    Q(processed=True, digr_approved=True)
                )
    all_approved = Files.objects.filter(send_to_sro = True,dept_approved=True,district_rgtr_approved=True,digr_approved=True)
    all_rejected = Files.objects.filter(
                    Q(processed=True, dept_approved=False) |
                    Q(processed=True, district_rgtr_approved=False) |
                    Q(processed=True, digr_approved=False)
                )
    
    # Count files by office (department-wise)

    office_counts = Files.objects.values('office__office_name').annotate(
        approved_count=Count('id', filter=Q(send_to_sro = True,dept_approved=True)),
        rejected_count=Count('id', filter=Q(send_to_sro = True,dept_approved=False)),
        pending_count=Count('id', filter=Q(send_to_sro = True,dept_approved=None)),
        total_count=Count('id', filter=Q(send_to_sro = True)),

    )
    # Data preparation for the office chart
    offices = list(Office.objects.values_list('office_name', flat=True))
    districts = list(District.objects.values_list('district_name', flat=True))
    zones = list(Zone.objects.values_list('zone_name', flat=True))

    # Count files by district registrar (district-wise)
    district_counts = Files.objects.values('office__district__district_name').annotate(
        approved_count=Count('id', filter=Q(send_to_sro = True,dept_approved=True,district_rgtr_approved=True)),
        rejected_count=Count('id', filter=Q(send_to_sro=True) & (
                                Q(dept_approved=False) | Q(district_rgtr_approved=False)
                            )),
        pending_count=Count('id', filter=Q(send_to_sro = True,dept_approved=True,district_rgtr_approved=None)),
        total_count_district=Count('id',)

    )
    

    # Count files by DIGR (zone-wise)
    zone_counts = Files.objects.values('office__district__zone__zone_name').annotate(
        approved_count=Count('id', filter=Q(send_to_sro = True,dept_approved=True,district_rgtr_approved=True,digr_approved=True)),
        rejected_count=Count('id', filter=Q(send_to_sro=True) & (
                                Q(dept_approved=False) | Q(district_rgtr_approved=False) | Q(digr_approved=False)
                            )),
        pending_count=Count('id', filter=Q(send_to_sro = True,dept_approved=True,district_rgtr_approved=True,digr_approved=None)),
        total_count_zone = Count('id', filter=Q(send_to_sro = True))
    )   
    
    office_data = []
    for office in offices:
        office_data.append({
            'name': office,
            'approved_count': next((o['approved_count'] for o in office_counts if o['office__office_name'] == office), 0),
            'rejected_count': next((o['rejected_count'] for o in office_counts if o['office__office_name'] == office), 0),
            'pending_count': next((o['pending_count'] for o in office_counts if o['office__office_name'] == office), 0),
            'total_count': next((o['total_count'] for o in office_counts if o['office__office_name'] == office), 0),
        })

    district_data = []
    for district in districts:
        district_data.append({
            'name': district,
            'approved_count': next((d['approved_count'] for d in district_counts if d['office__district__district_name'] == district), 0),
            'rejected_count': next((d['rejected_count'] for d in district_counts if d['office__district__district_name'] == district), 0),
            'pending_count': next((d['pending_count'] for d in district_counts if d['office__district__district_name'] == district), 0),
            'total_count_district': next((d['total_count_district'] for d in district_counts if d['office__district__district_name'] == district), 0),

        })

    zone_data = []
    for zone in zones:
        zone_data.append({
            'name': zone,
            'total_count_zone': next((z['total_count_zone'] for z in zone_counts if z['office__district__zone__zone_name'] == zone), 0),
            'approved_count': next((z['approved_count'] for z in zone_counts if z['office__district__zone__zone_name'] == zone), 0),
            'rejected_count': next((z['rejected_count'] for z in zone_counts if z['office__district__zone__zone_name'] == zone), 0),
            'pending_count': next((z['pending_count'] for z in zone_counts if z['office__district__zone__zone_name'] == zone), 0),
        })
        

    total_pages = 0
    processed_files = Files.objects.filter(processed=True)

    total_pages = sum(file.page_count or 0 for file in processed_files)
    
    received_documents = ReceivedDocuments.objects.all()
    total_received = received_documents.aggregate(total=Sum('received'))['total'] or 0
    in_process = (all_submitted.count() - all_pending.count()) - all_rejected.count()

    approved_by_sro = max(0, Files.objects.filter(processed=True, dept_approved=True).count() - 87)
    rejected_by_sro = Files.objects.filter(processed = True,dept_approved = False)
    pending_by_sro = Files.objects.filter(processed = True,dept_approved = None)

    approved_by_dr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True)
    rejected_by_dr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = False)
    pending_by_dr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = None)

    approved_by_digr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True, digr_approved =True)
    rejected_by_digr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True, digr_approved =False)
    pending_by_digr = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True, digr_approved =None)

    context = {
        'office_data': office_data,
        'district_data': district_data,
        'zone_data': zone_data,
        'offices': offices,
        'districts': districts,
        'zones': zones,
        'all_submitted':all_submitted,
        'all_approved':all_approved,
        'all_rejected':all_rejected,
        'total_pages':total_pages,
        'total_received':total_received,
        'in_process':in_process,
        'all_pending':all_pending,
        'pending_by_sro':pending_by_sro,
        'approved_by_sro':approved_by_sro,
        'rejected_by_sro':rejected_by_sro,
        'pending_by_dr':pending_by_dr,
        'approved_by_dr':approved_by_dr,
        'rejected_by_dr':rejected_by_dr,
        'pending_by_digr':pending_by_digr,
        'approved_by_digr':approved_by_digr,
        'rejected_by_digr':rejected_by_digr

    }
    return render(request, 'users/igr_dashboard.html', context)



def agencies(request):
    agencies = Agency.objects.all()
    context = {'agencies': agencies}
    return render(request, 'users/agencies.html', context)


def filtered_data_view(request):
    document_type = request.GET.getlist("document_type", [])
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    queryset = Files.objects.all()

    if document_type:
        queryset = queryset.filter(filename__in=document_type)

    if from_date:
        queryset = queryset.filter(uploaded_at__gte=from_date)

    if to_date:
        queryset = queryset.filter(uploaded_at__lte=to_date)

    context = {"documents": queryset}

    return render(request, 'users/filtered_data.html', context)


@login_required
def profile_user(request):
    user = request.user

    if request.method == 'POST':
        current_password = request.POST.get('password')
        new_password = request.POST.get('newpassword')
        confirm_new_password = request.POST.get('renewpassword')

        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')

            return render(request, 'users/profile.html', {'user': user})

        if new_password != confirm_new_password:
            messages.error(request, 'New password and confirm new password do not match.')

            return render(request, 'users/profile.html', {'user': user})

        if new_password:
            user.set_password(new_password)
            user.save()
            
            update_session_auth_hash(request, user)
            
            messages.success(request, 'Password changed successfully!')
            return render(request, 'users/profile.html', {'user': user})

    return render(request, 'users/profile.html', {'user': user})

@login_required
def users(request):
    
    if request.user.is_department:
        return redirect('dashboard')
    
    
    search_query = request.GET.get('q', '')
    users = User.objects.all().order_by('id')
    
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(contact__icontains=search_query) | 
            Q(email__icontains=search_query) | 
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) 
        )
    paginator = Paginator(users, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'users/users.html', {
        'users': page_obj.object_list,
        'total':users,
        'page_obj': page_obj,
        'search_query': search_query,
    })
    


@login_required
def create_user_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.is_department:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        contact = request.POST.get('contact')
        office = request.POST.get('office')  
        is_department = request.POST.get('is_department') == 'on'
        is_district_rgtr = request.POST.get('is_district_rgtr') == 'on'
        is_digr = request.POST.get('is_digr') == 'on'
        is_igr = request.POST.get('is_igr') == 'on'
        is_agency_admin = request.POST.get('is_agency_admin') == 'on'       
        
        district_id = request.POST.get('district_id')
        zone_id = request.POST.get('zone_id')
        agency_id = request.POST.get('agency_id')
        
        if is_agency_admin and not agency_id:
            messages.error(request, "Agency is required when 'Agency Admin' is selected.")
            return redirect('create_user')
        
        if is_district_rgtr and not district_id:
            messages.error(request, "District is required when 'Is District Admin' is selected.")
            return redirect('create_user')

        if is_digr and not zone_id:
            messages.error(request, "Zone is required when 'Is Deputy IGR' is selected.")
            return redirect('create_user')

        district_instance = District.objects.get(id=district_id) if district_id else None
        zone_instance = Zone.objects.get(id=zone_id) if zone_id else None
        
        agency_instance = Agency.objects.get(id=agency_id) if agency_id else None
       
        office_instance = Office.objects.get(id=office) if office else None

        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect('create_user')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('create_user')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect('create_user')

        user = User.objects.create_user(username=username,
                                        first_name=first_name,
                                        last_name=last_name,
                                        office=office_instance,
                                        is_digr=is_digr,
                                        is_district_rgtr=is_district_rgtr,
                                        email=email,
                                        district=district_instance,
                                        zone=zone_instance,
                                        is_department=is_department,
                                        contact=contact,
                                        is_igr=is_igr,
                                        is_agency_admin=is_agency_admin,
                                        agency = agency_instance,
                                        password=password1)

        avatar = request.FILES.get('avatar')
        if avatar:
            user.avatar = avatar

        user.save()

        messages.success(request, "User created successfully.")
        return redirect('users')     

    offices = Office.objects.all()
    zones = Zone.objects.all()
    districts = District.objects.all()
    agencies = Agency.objects.all()
    
    context = {'offices': offices, 'zones': zones, 'districts': districts,'agencies':agencies}
    return render(request, 'users/create_user.html', context)


@login_required
def create_agency_user_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.is_department:
        return redirect('dashboard')

    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('email')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        office = request.POST.get('office') 
        agency = request.user.agency
        is_agency_qc_employee = request.POST.get('is_agency_qc_employee') == 'on' 
        is_agency_scanning_employee = request.POST.get('is_agency_scanning_employee') == 'on' 

        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect('create_agency_user')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('create_agency_user')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect('create_agency_user')
        
        office_instance = Office.objects.get(id=office) if office else None

        user = User.objects.create_user(username=username,
                                        first_name=first_name,
                                        last_name=last_name,
                                        office=office_instance,
                                        email=email,
                                        contact=contact,
                                        agency=agency,
                                        is_agency_qc_employee = is_agency_qc_employee ,
                                        is_agency_scanning_employee = is_agency_scanning_employee,
                                        password=password1)

        avatar = request.FILES.get('avatar')
        if avatar:
            user.avatar = avatar

        user.save()

        messages.success(request, "User created successfully.")
        return redirect('agency_users')     

    offices = Office.objects.all()
    context = {
        'offices' :offices ,
    }
    return render(request, 'users/create_agency_user.html',context)

@login_required
def create_agencies_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.is_department:
        return redirect('dashboard')

    if request.method == 'POST':
        name = request.POST.get('name')
        address = request.POST.get('address')
        contact_number = request.POST.get('contact_number')
        office = request.POST.get('office') 
        office_instance = Office.objects.get(id=office) if office else None
        if Agency.objects.filter(name=name).exists():
            messages.error(request, "Agency already exists.")
            return redirect('create_agency')


        agency = Agency.objects.create( name = name,
                                        address=address,
                                        contact_number=contact_number,
                                        office=office_instance,
                                    )

        messages.success(request, "Agency created successfully.")
        return redirect('agencies')     

    offices = Office.objects.all()
    
    context = {'offices': offices}
    return render(request, 'users/create_agency.html', context)



@login_required
def update_user_view(request, user_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.is_department:
        return redirect('dashboard')

    if not request.user.is_admin:
        messages.error(request, "You do not have permission to update users.")
        return redirect('dashboard')

    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        office_code = request.POST.get('office_code')
        email = request.POST.get('email')
        contact = request.POST.get('contact')
        is_active = request.POST.get('is_active') == 'on'
            

        user.first_name = first_name
        user.last_name = last_name
        user.office_code = office_code
        user.email = email
        user.contact = contact
        
        user.is_active = is_active
        user.save()
        if user.is_department:
            
            messages.success(request, f"User {user.username} updated successfully!")
            return redirect('dept_users')


        messages.success(request, f"User {user.username} updated successfully!")
        return redirect('users')

    context = {
        'user': user,
    }

    return render(request, 'users/user_profile.html', context)

def agency_users(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.is_department:
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    users = User.objects.filter(agency = request.user.agency)
    
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(contact__icontains=search_query) | 
            Q(email__icontains=search_query) | 
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) 
        )
    paginator = Paginator(users, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'users': page_obj.object_list,
        'page_obj': page_obj,
        'search_query': search_query,
        'total':users
        }
    
    return render(request, 'users/agency_users.html', context)



def office_wise_page_report(request):
    office_data = {}
    processed_files = Files.objects.filter(processed=True)
    
    for file in processed_files:
        office_name = file.office.office_name if file.office else "Unknown"
        file_path = Path(file.processed_file.path)
        page_count = count_tiff_pages(file_path)
        
        if office_name not in office_data:
            office_data[office_name] = {"files": 0, "pages": 0}
        
        office_data[office_name]["files"] += 1
        office_data[office_name]["pages"] += page_count
    
    df = pd.DataFrame.from_dict(office_data, orient='index')
    df.reset_index(inplace=True)
    df.rename(columns={"index": "Office", "files": "Files", "pages": "Pages"}, inplace=True)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=office_wise_report.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Office Wise Report')
    
    return response


def office_wise_page_report(request):
    office_data = {}
    processed_files = Files.objects.filter(processed=True)
    
    for file in processed_files:
        office_name = file.office.office_name if file.office else "Unknown"
        page_count = file.page_count or 0
        
        if office_name not in office_data:
            office_data[office_name] = {"files": 0, "pages": 0}
        
        office_data[office_name]["files"] += 1
        office_data[office_name]["pages"] += page_count
    
    df = pd.DataFrame.from_dict(office_data, orient='index')
    df.reset_index(inplace=True)
    df.rename(columns={"index": "Office", "files": "Files", "pages": "Pages"}, inplace=True)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=office_wise_report.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Office Wise Report')
    
    return response


from django.http import JsonResponse
def update_existing_files_page_count(request):
    files = Files.objects.all()
    
    success_count = 0
    error_count = 0
    error_files = []

    for file_obj in files:
        if file_obj.processed_file:
            try:
                file_obj.processed_file.open('rb')
                with Image.open(file_obj.processed_file) as img:
                    page_count = getattr(img, 'n_frames', 1)
                    file_obj.page_count = page_count
                    file_obj.save()
                    success_count += 1
                    print(f"Updated {file_obj.filename}: {page_count} pages")
            except Exception as e:
                print(f"Error processing {file_obj.filename}: {e}")
                file_obj.page_count = 1
                file_obj.save()
                error_count += 1
                error_files.append(file_obj.filename)
            finally:
                file_obj.processed_file.close()
    
    response_data = {
        'success_count': success_count,
        'error_count': error_count,
        'error_files': error_files,
    }

    return JsonResponse(response_data)



@login_required
def activities_get(request):
    activities = UserActivity.objects.all()
    users = User.objects.all()

    search_query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    username = request.GET.get('username', '')

    if search_query:
        activities = activities.filter(user__username__icontains=search_query)

    if username:
        activities = activities.filter(user__username=username)

    if start_date:
        activities = activities.filter(timestamp__date__gte=start_date)
    if end_date:
        activities = activities.filter(timestamp__date__lte=end_date)

    paginator = Paginator(activities, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "activities": page_obj,
        "search_query": search_query,
        "start_date": start_date,
        "end_date": end_date,
        "username": username,
        "users": users,
    }
    return render(request, 'users/activities.html', context)


@login_required
def download_activities_excel(request):
    search_query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    username = request.GET.get('username', '')

    activities = UserActivity.objects.all()

    if search_query:
        activities = activities.filter(user__username__icontains=search_query)

    if username:
        activities = activities.filter(user__username=username)

    if start_date:
        activities = activities.filter(timestamp__date__gte=start_date)
    if end_date:
        activities = activities.filter(timestamp__date__lte=end_date)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "User Activities"

    headers = ['User', 'Office', 'Action', 'Date and Time', 'IP Address', 'Device Type', 'Browser', 'OS']
    worksheet.append(headers)

    for activity in activities:
        worksheet.append([
            activity.user.username,
            activity.user.office.office_name if activity.user.office else '',
            activity.action,
            activity.formatted_timestamp,
            activity.ip_address,
            activity.device_type,
            activity.browser,
            activity.os,
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="user_activities.xlsx"'

    return response


@login_required
def agency_wise_report(request, agency_id):
    agency = Agency.objects.get(id=agency_id)
    search_query = request.GET.get('q', '').strip()

    users_qs = User.objects.filter(agency=agency)

    if search_query:
        users_qs = users_qs.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )

    users = users_qs.annotate(
                                total_uploaded=Count('uploaded_files'),
                                dept_approved_count=Count(
                                    Case(
                                        When(uploaded_files__dept_approved=True, then=1),
                                        output_field=IntegerField(),
                                    )
                                ),
                                dr_approved_count=Count(
                                    Case(
                                        When(uploaded_files__district_rgtr_approved=True, then=1),
                                        output_field=IntegerField(),
                                    )
                                ),
                                digr_approved_count=Count(
                                    Case(
                                        When(uploaded_files__digr_approved=True, then=1),
                                        output_field=IntegerField(),
                                    )
                                ),
                                rejected_count=Count(
                                    Case(
                                        When(
                                            Q(uploaded_files__dept_approved=False) |
                                            Q(uploaded_files__district_rgtr_approved=False) |
                                            Q(uploaded_files__admin_approved=False) |
                                            Q(uploaded_files__digr_approved=False),
                                            then=1
                                        ),
                                        output_field=IntegerField(),
                                    )
                                ),
                                total_page_count=Sum('uploaded_files__page_count'),
                            ).filter(total_uploaded__gt=0)

    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    files_qs = Files.objects.filter(uploaded_by__in=users_qs)

    totals = files_qs.aggregate(
        total_uploaded=Count('id'),
        dept_approved_total=Count('id', filter=Q(dept_approved=True)),
        dr_approved_total=Count('id', filter=Q(district_rgtr_approved=True)),
        digr_approved_total=Count('id', filter=Q(digr_approved=True)),

        rejected_total=Count('id', filter=Q(dept_approved=False) | Q(district_rgtr_approved=False) | Q(digr_approved=False)),

        pending_total=Count('id', filter=Q(dept_approved__isnull=True) | Q(district_rgtr_approved__isnull=True) | Q(digr_approved__isnull=True)),

        total_page_count=Sum('page_count'),
    )

    context = {
        'users_data': page_obj,
        'page_obj': page_obj,
        'agency': agency,
        'search_query': search_query,
        'total': users,
        'totals': totals,
    }
    return render(request, 'users/agency_wise_report.html', context)

def export_agency_report_excel(request, agency_id):
    agency = Agency.objects.get(id=agency_id)
    search_query = request.GET.get('q', '').strip()

    users_qs = User.objects.filter(agency=agency)

    if search_query:
        users_qs = users_qs.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )

    users = users_qs.annotate(
        total_uploaded=Count('uploaded_files'),
        dept_approved_count=Count('uploaded_files', filter=Q(uploaded_files__dept_approved=True)),
        dr_approved_count=Count('uploaded_files', filter=Q(uploaded_files__district_rgtr_approved=True)),
        digr_approved_count=Count('uploaded_files', filter=Q(uploaded_files__digr_approved=True)),
        rejected_count=Count(
            'uploaded_files',
            filter=Q(uploaded_files__dept_approved=False) |
                   Q(uploaded_files__district_rgtr_approved=False) |
                   Q(uploaded_files__digr_approved=False)
        ),
        total_page_count=Sum('uploaded_files__page_count'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{agency.name} Report"

    headers = [
        'Full Name', 'Username', 'Total Uploaded',
        'Dept Approved', 'DR Approved', 'DIGR Approved',
        'Rejected Files', 'Total Page Count'
    ]
    ws.append(headers)

    for user in users:
        ws.append([
            user.get_full_name(),
            user.username,
            user.total_uploaded,
            user.dept_approved_count,
            user.dr_approved_count,
            user.digr_approved_count,
            user.rejected_count,
            user.total_page_count or 0
        ])

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{agency.name}_report.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

FALLBACK_JSON_PATH = Path(__file__).resolve().parent / "fallback_data.json"


def get_external_dashboard_data():
    try:
        response = requests.get("https://docdataentryapis.shaleemardemo.com/api/CronJob/DashboardInfo", timeout=5)
        if response.status_code == 200:
            data = response.json()

            # ✅ Optionally save this to fallback file for future use
            with open(FALLBACK_JSON_PATH, "w") as f:
                json.dump(data, f)

            return data
    except requests.RequestException:
        pass  # Continue to fallback

    try:
        with open(FALLBACK_JSON_PATH, "r") as f:
            return json.load(f)
    except Exception:
        return {
            "firstDECounts": 0,
            "secondDECounts": 0,
            "qcdeCounts": 0,
            "totalSROReceived": 0,
            "totalSROApproval": 0,
            "totalSRORejected": 0,
            "fileInfos": [],
        }


def reports_dasboards(request):
    filters = Q(send_to_sro=True, office__isnull=False)

    sro_report = Files.objects.filter(filters).values(
        'office__office_name'
    ).annotate(
        total_qc=Count('id'),
        approved_count=Count('id', filter=Q(dept_approved=True)),
        rejected_count=Count('id', filter=Q(dept_approved=False)),
        pending_count=Count('id', filter=Q(dept_approved=None)),
        send_to_qc_pages=Sum('page_count'),
        approved_pages=Sum('page_count', filter=Q(dept_approved=True)),
        rejected_pages=Sum('page_count', filter=Q(dept_approved=False)),
        pending_pages=Sum('page_count', filter=Q(dept_approved=None)),
    ).order_by('office__office_name')

    sro_data = [
        {
            "s_no": index,
            "sro_name": item['office__office_name'],
            "total_qc": item['total_qc'],
            "approved": item['approved_count'],
            "rejected": item['rejected_count'],
            "pending": item['pending_count'],
            "approved_pages": item.get("approved_pages") or 0,
            "rejected_pages": item.get("rejected_pages") or 0,
            "pending_pages": item.get("pending_pages") or 0,
            "send_to_qc_pages": item.get("send_to_qc_pages") or 0,
        }
        for index, item in enumerate(sro_report, start=1)
    ]

    sro_count = len(sro_data)
    total_approved_pages = sum(item["approved_pages"] for item in sro_data)
    total_send_to_qc_pages = sum(item["send_to_qc_pages"] for item in sro_data)

    total_pages = Files.objects.filter(processed=True).aggregate(
        total=Sum('page_count')
    )['total'] or 0

    three_days_ago = timezone.now() - timedelta(days=1)
    active_offices = Files.objects.filter(
        uploaded_at__gte=three_days_ago,
        office__isnull=False
    ).values_list('office__office_name', flat=True).distinct()

    active_sro_count = active_offices.count()
    active_sro_names = list(active_offices)

    # ✅ Fetch external API data
    external_data = get_external_dashboard_data()



    context = {
        "sro_count": sro_count,
        "active_sro_count": active_sro_count,
        "active_sro_names": active_sro_names,
        "total_approved_pages": total_approved_pages,
        "total_send_to_qc_pages": total_send_to_qc_pages,
        "total_pages": total_pages,

        # ✅ External API data
        "firstDECounts": external_data.get("firstDECounts", 0),
        "secondDECounts": external_data.get("secondDECounts", 0),
        "qcdeCounts": external_data.get("qcdeCounts", 0),
        "totalSROReceived": external_data.get("totalSROReceived", 0),
        "totalSROApproval": external_data.get("totalSROApproval", 0),
        "totalSRORejected": external_data.get("totalSRORejected", 0),
        "fileInfos": external_data.get("fileInfos", []),
    }

    return render(request, 'users/all_reports_dashboard.html', context)
