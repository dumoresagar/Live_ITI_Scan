from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Files,ReceivedDocuments
from django.contrib import messages
from django.http import FileResponse,Http404, HttpResponse
import zipfile
from io import BytesIO
from django.core.paginator import Paginator
from django.utils import timezone
import os
from django.core.files.storage import default_storage
from itertools import groupby
from operator import attrgetter
from users.models import UserActivity
from django.db.models import Q
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.db.models import Count, Sum, F
from django.db.models.functions import TruncDate
from users.models import Office, User
import re
from openpyxl import Workbook
from django.views.decorators.http import require_POST
import logging
from .serializer import FilesUploadSerializer
logger = logging.getLogger(__name__)
from django.shortcuts import get_list_or_404
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from .serializer import FilesUploadSerializer
import tempfile
from django.utils.dateparse import parse_date
from rest_framework.permissions import AllowAny
from rest_framework.authentication import TokenAuthentication
from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response

def log_user_activity(request):
    if request.user.is_authenticated:
        logger.info(f"User {request.user.username} performed an action.")



# Create your views here.

OFFICE_CODE_PATTERNS = [
    r"^I_(\d{1,3})_",        # Index File
    r"^MTPR_(\d{1,3})_",     # Municipal Town Property Register
    r"^RH_(\d{1,3})_",       # Register of Holdings
    r"^R_(\d{1,3})_",        # Regular Document Register
    r"^LO_(\d{1,3})_",       # Loan Order Register
    r"^MO_(\d{1,3})_",       # Memo Order Register
    r"^CO_(\d{1,3})_"        # Court Order Register
]

def extract_office_code(filename):
    """Extracts the Office Code from the filename using regex patterns."""
    for pattern in OFFICE_CODE_PATTERNS:
        match = re.match(pattern, filename)
        if match:
            return match.group(1).zfill(3)  # Ensure 3-digit format
    return None


@login_required
def upload_processed_files(request):
    upload_count = 0
    replace_count = 0
    replaced_files = []  

    if request.method == 'POST' and request.FILES.getlist('processed_files'):
        files = request.FILES.getlist('processed_files')
        date = request.POST.get('date')

        if date:
            date = datetime.strptime(date, "%Y-%m-%d")
        else:
            date = timezone.now()
            
        files_without_office = []

        for file in files:

            filename = os.path.splitext(file.name)[0]
            office_code = extract_office_code(filename)
            office = Office.objects.filter(office_code=office_code).first() if office_code else None

            if office is None:
                files_without_office.append(filename)
                continue  # Skip processing this file

            existing_file = Files.objects.filter(filename=filename).first()
            
            if request.user.is_admin:
                if existing_file:
                    if existing_file.processed_file:
                        file_path = existing_file.processed_file.path
                        if default_storage.exists(file_path):
                            default_storage.delete(file_path)

                    existing_file.processed_file = file
                    existing_file.uploaded_at = date
                    existing_file.uploaded_by = request.user
                    existing_file.dept_approved = None
                    existing_file.processed = True
                    existing_file.admin_approved = True
                    existing_file.send_to_sro = True 
                    existing_file.district_rgtr_approved = None
                    existing_file.digr_approved = None

                    if existing_file.office is None and office:
                        existing_file.office = office

                    existing_file.save()

                    UserActivity.objects.create(
                        user=request.user,
                        action=f"Replaced file '{filename}'",
                        category='info',
                        ip_address=request.ip_address,  
                        device_type=request.device_type,  
                        browser=request.browser,  
                        os=request.os 
                    )
                    replace_count += 1
                    replaced_files.append(filename)

                else:
                    processed_file_instance = Files(
                        processed_file=file,
                        filename=filename,
                        uploaded_by=request.user,
                        processed=True,
                        uploaded_at=date,
                        office=office
                    )
                    processed_file_instance.save()

                    UserActivity.objects.create(
                        user=request.user,
                        action=f"Uploaded new file '{filename}'",
                        category='success',
                        ip_address=request.ip_address,  
                        device_type=request.device_type,  
                        browser=request.browser,  
                        os=request.os 
                    )
                    upload_count += 1
            else:
                if existing_file:
                    if existing_file.processed_file:
                        file_path = existing_file.processed_file.path
                        if default_storage.exists(file_path):
                            default_storage.delete(file_path)

                    existing_file.processed_file = file
                    existing_file.uploaded_at = timezone.now()
                    existing_file.uploaded_by = request.user
                    existing_file.admin_approved = None
                    existing_file.dept_approved = None
                    existing_file.processed = True
                    existing_file.send_to_qc = True 
                    existing_file.district_rgtr_approved = None
                    existing_file.digr_approved = None

                    if existing_file.office is None and office:
                        existing_file.office = office

                    existing_file.save()

                    UserActivity.objects.create(
                        user=request.user,
                        action=f"Replaced file '{filename}'",
                        category='info',
                        ip_address=request.ip_address,  
                        device_type=request.device_type,  
                        browser=request.browser,  
                        os=request.os 
                    )
                    replace_count += 1
                    replaced_files.append(filename)

                else:
                    processed_file_instance = Files(
                        processed_file=file,
                        filename=filename,
                        uploaded_by=request.user,
                        processed=True,
                        uploaded_at=timezone.now(),
                        office=office,
                    )
                    processed_file_instance.save()


                    UserActivity.objects.create(
                        user=request.user,
                        action=f"Uploaded new file '{filename}'",
                        category='success',
                        ip_address=request.ip_address,  
                        device_type=request.device_type,  
                        browser=request.browser,  
                        os=request.os 
                    )
                    upload_count += 1

        if files_without_office:
            missing_offices_str = ", ".join(files_without_office)
            messages.error(
                request,
                f"The following files were not uploaded because no office was found: {missing_offices_str}"
            )

        if replaced_files:
            replaced_files_str = ", ".join(replaced_files)
            messages.info(
                request, 
                f"The following files were replaced successfully: {replaced_files_str}"
            )
        elif not upload_count:
            messages.info(request, "No new files were uploaded.")

        messages.success(request, "File processing completed.")
        if request.user.is_admin:
            return redirect('send_to_sro')
        else:
            return redirect('send_to_qc')

    return render(request, 'documents/upload_processed.html')

@login_required
def send_to_sro(request):
    files = Files.objects.filter(uploaded_by=request.user, send_to_sro=False)
    
    if request.method == "POST":
        processed_count = 0  # Counter for successfully processed files
        
        for file in files:
            query = request.POST.get(f'query_{file.id}', '').strip()  # Get and sanitize query
            if query:
                file.remark_by_scanner = query[:50]  # Truncate to 50 characters
            else:
                file.remark_by_scanner = None  # Set to None if no query provided
            file.send_to_sro = True
            file.save()
            processed_count += 1
        
        # Success message
        messages.success(request, f"{processed_count} file(s) uploaded successfully")
        return redirect('processed_files')
    
    # Render page with file list
    return render(request, 'documents/send_to_sro.html', {'files': files})

@login_required
def send_to_qc(request):
    files = Files.objects.filter(uploaded_by=request.user, send_to_qc=False)
    
    if request.method == "POST":
        processed_count = 0  # Counter for successfully processed files
        
        for file in files:
            query = request.POST.get(f'query_{file.id}', '').strip()  # Get and sanitize query
            if query:
                file.remark_by_scanner = query[:50]  # Truncate to 50 characters
            else:
                file.remark_by_scanner = None  # Set to None if no query provided
            file.send_to_qc = True
            file.save()
            processed_count += 1
        
        # Success message
        messages.success(request, f"{processed_count} file(s) uploaded successfully")
        return redirect('processed_files')
    
    # Render page with file list
    return render(request, 'documents/send_to_qc.html', {'files': files})
@login_required
def upload_unprocessed_files(request):
    if request.user.is_department:
        return redirect('dashboard')
    
    
    if request.method == 'POST' and request.FILES.getlist('unprocessed_files'):
        uploaded_files = request.FILES.getlist('unprocessed_files')
        
        for file in uploaded_files:
            unprocessed_file_instance = Files(processed_file=file)
            filename = unprocessed_file_instance.extract_filename(file.name)
            
            if Files.objects.filter(filename=filename).exists():
                messages.info(request, f"File '{filename}' already exists and was skipped.")
                continue 
            
            unprocessed_file_instance.filename = filename
            unprocessed_file_instance.uploaded_by = request.user
            unprocessed_file_instance.processed = False
            unprocessed_file_instance.save()
        
        return redirect('processed_files') 
        
    return render(request, 'documents/upload_unprocessed.html')


@login_required
def processed_files(request):
    if request.user.is_department:
        return redirect('dashboard')
    
    
    search_query = request.GET.get('q', '')  
    
    if request.user.is_agency_qc_employee or request.user.is_agency_admin:
        print("yes")
        files = Files.objects.filter(uploaded_by = request.user ,processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at')
    else:
        files = Files.objects.filter(processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at')


    if search_query:
        files = files.filter(filename__icontains=search_query)  

    paginator = Paginator(files, 10)  
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number) 

    return render(request, 'documents/processed_files.html', {
        'files':files,
        'page_obj': page_obj,
        'search_query': search_query, 
    })



@login_required
def processed_files_self_agency(request):
    if request.user.is_department:
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')  

    if search_query:
        files = files.filter(filename__icontains=search_query)  

    paginator = Paginator(files, 10)  
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number) 

    return render(request, 'documents/processed_files_selfagencies.html', {
        'files':files,
        'page_obj': page_obj,
        'search_query': search_query, 
    })

@login_required
def processed_file(request, file_id):
    file = Files.objects.get(id=file_id) 
    if request.method == 'POST':
        if request.user.is_department:
            remark = request.POST.get('remark')
            
            remark_with_username = f"({request.user.username}) {remark} "

            file.remark = remark_with_username
            file.dept_approved = False
            file.dept_approved_by = request.user
            file.dept_approved_at = timezone.now()  
            file.save()
            
            UserActivity.objects.create(
                    user=request.user,
                    action=f"Rejected document '{file.filename}'",
                    category='danger',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )

            messages.success(request, f"{file.filename} has been Rejected.")
            
            # Get the next file
            next_file = Files.objects.filter(office=request.user.office,processed=True,send_to_sro = True, dept_approved=None, district_rgtr_approved=None,digr_approved=None).order_by('-uploaded_at').first()
            
            if next_file:
                return redirect('processed_file', file_id=next_file.id)  # Redirect to the next file
            else:
                return redirect('files_for_qc')  # If no next file, redirect to the approved files list

        else:
            messages.error(request, "You are not authorized to perform this action.")
            return redirect('files_for_qc')
        
    grouped_files = Files.objects.filter(office=request.user.office,processed=True, dept_approved=None).order_by('uploaded_by', '-uploaded_at')

    index_files = {}
    for user, files in groupby(grouped_files, key=attrgetter('uploaded_by')):
        index_files[user] = list(files)

    
    return render(request, 'documents/processed_file.html', {'file': file,'index_files':index_files})


     
@login_required
def approved_file(request, file_id):
    file = Files.objects.get(id = file_id)
    index_files = Files.objects.filter(processed = True, dept_approved = True).order_by('-dept_approved_at') 
    return render(request, 'documents/approved_file.html', {'file': file,'index_files':index_files})


@login_required
def rejected_file(request, file_id):
    file = Files.objects.get(id = file_id) 
    if request.user.is_department:
        index_files = (
                            Files.objects.filter(
                                Q(office = request.user.office,processed=True, dept_approved=False) |
                                Q(office = request.user.office,processed=True, district_rgtr_approved=False) |
                                Q(office = request.user.office,processed=True, digr_approved=False)
                            )
                            .order_by('uploaded_at')
                        ) 
    elif request.user.is_district_rgtr:
        index_files = (
                            Files.objects.filter(
                                Q(office__district = request.user.district,processed=True, dept_approved=False) |
                                Q(office__district = request.user.district,processed=True, district_rgtr_approved=False) |
                                Q(office__district = request.user.district,processed=True, digr_approved=False)
                            )
                            .order_by('uploaded_at')
                        )
    else:
        index_files = (
                            Files.objects.filter(
                                Q(processed=True, dept_approved=False) |
                                Q(processed=True, district_rgtr_approved=False) |
                                Q(processed=True, digr_approved=False)
                            )
                            .order_by('uploaded_at')
                        )  
        
    return render(request, 'documents/rejected_file.html', {'file': file,'index_files':index_files})


def approve_rejected_file(request, file_id):
    file = Files.objects.get(id = file_id) 
    if request.user.is_department: # is_district_rgtr is_digr
        file.dept_approved = True
        file.save()
    elif request.user.is_district_rgtr:
        file.dept_approved = True    
        file.district_rgtr_approved = True
        file.save()
    elif request.user.is_digr:
        file.dept_approved = True    
        file.district_rgtr_approved = True
        file.digr_approved = True
        file.save()
        
    next_file= (
                    Files.objects.filter(
                        Q(processed=True, dept_approved=False) |
                        Q(processed=True, district_rgtr_approved=False) |
                        Q(processed=True, digr_approved=False)
                    )
                    .order_by('uploaded_at')
                ).first()
    if next_file:
        return redirect('rejected_file', file_id = next_file.id)
    else:
        return redirect('rejected_files')

@login_required
def download_file(request, file_id):
    try:
        file = Files.objects.get(id=file_id)
    except Files.DoesNotExist:
        raise Http404("File not found")

    return FileResponse(file.processed_file, as_attachment=True)
    
@login_required
def scanned_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(processed = False) 
    if search_query:
        files = files.filter(filename__icontains=search_query)

    paginator = Paginator(files, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/scanned_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })


@login_required
def approved_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved= True,
                                 digr_approved= True,dept_approved_by= request.user).order_by('-dept_approved_at')  
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/approved_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })

@login_required
def sro_approved_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(office__district = request.user.district,processed = True, dept_approved = True,district_rgtr_approved = None).order_by('uploaded_at')  
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/sro_approved_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })


@login_required
def sro_approved_file(request, file_id):
    file = Files.objects.get(id=file_id) 
    if request.method == 'POST':
        if request.user.is_district_rgtr:
            remark = request.POST.get('remark')
            remark_with_username = f"({request.user.username}) {remark} "
            
            file.remark = remark_with_username
            file.district_rgtr_approved = False
            file.district_rgtr_approved_by = request.user
            file.district_rgtr_approved_at = timezone.now() 
            file.save()
            
            
            UserActivity.objects.create(
                    user=request.user,
                    action=f"Rejected document '{file.filename}'",
                    category='danger',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )

            messages.success(request, f"{file.filename} has been Rejected.")
            
            # Get the next file
            next_file = Files.objects.filter(office__district = request.user.district,processed=True, dept_approved=True,district_rgtr_approved=None,digr_approved=None).order_by('-dept_approved_at').first()
            
            if next_file:
                return redirect('sro_approved_file', file_id=next_file.id)  # Redirect to the next file
            else:
                return redirect('sro_approved_files')  # If no next file, redirect to the approved files list

        else:
            messages.error(request, "You are not authorized to perform this action.")
            return redirect('sro_approved_files')
        
    grouped_files = Files.objects.filter(office__district = request.user.district,processed = True, dept_approved = True,district_rgtr_approved = None).order_by('dept_approved_by','-dept_approved_at')

    index_files = {}
    for user, files in groupby(grouped_files, key=attrgetter('dept_approved_by')):
        index_files[user] = list(files)
        
    return render(request, 'documents/sro_approved_file.html', {'file': file, 'index_files':index_files})


@login_required
def d_a_approved_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None).order_by('uploaded_at')  
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/d_a_approved_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })

@login_required
def d_a_approved_file(request, file_id):
    file = Files.objects.get(id=file_id) 
    
    if request.method == 'POST':
        if request.user.is_digr:
            remark = request.POST.get('remark')
            remark_with_username = f"({request.user.username}) {remark} "
            file.remark = remark_with_username
            file.digr_approved = False
            file.digr_approved_by = request.user
            file.digr_approved_at = timezone.now() 
            file.save()
            messages.success(request, f"{file.filename} has been Rejected.")
            
            UserActivity.objects.create(
                    user=request.user,
                    action=f"Rejected document '{file.filename}'",
                    category='danger',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )
            
            # Get the next file
            next_file = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None).order_by('-district_rgtr_approved_at').first()
            
            if next_file:
                return redirect('d_a_approved_file', file_id=next_file.id)  # Redirect to the next file
            else:
                return redirect('d_a_approved_files')  # If no next file, redirect to the approved files list

        else:
            messages.error(request, "You are not authorized to perform this action.")
            return redirect('d_a_approved_files')
    grouped_files = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None).order_by('district_rgtr_approved_by','uploaded_at')

    index_files = {}
    for user, files in groupby(grouped_files, key=attrgetter('district_rgtr_approved_by')):
        index_files[user] = list(files) 
        
    return render(request, 'documents/d_a_approved_file.html', {'file': file,'index_files':index_files})

@login_required
def d_a_approved_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(office__district__zone = request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None).order_by('uploaded_at')  
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/d_a_approved_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })




@login_required
def rejected_files(request):
    search_query = request.GET.get('q', '')

    if request.user.is_admin or request.user.is_igr:
    
        files = (
                    Files.objects.filter(
                        Q(processed=True, dept_approved=False) |
                        Q(processed=True, district_rgtr_approved=False) |
                        Q(processed=True, digr_approved=False)
                    )
                    .order_by(
                                '-dept_approved_at', '-district_rgtr_approved_at', '-digr_approved_at'
                            )
                )
    elif request.user.is_department:
        files = (
                    Files.objects.filter(
                        Q(office = request.user.office,processed=True, dept_approved=False)
                        
                    )
                    .order_by(
                                '-dept_approved_at',
                            )
                )
    elif request.user.is_district_rgtr:
        files = (
                    Files.objects.filter(
                        Q(office__district = request.user.district,processed=True, district_rgtr_approved=False)
                        
                    )
                    .order_by(
                                '-district_rgtr_approved_at',
                            )
                )
    elif request.user.is_igr:
        files = (
                    Files.objects.filter(
                        Q(office__district__zone = request.user.zone,processed=True, digr_approved=False)
                        
                    )
                    .order_by(
                                '-digr_approved_at',
                            )
                )

    
    
    
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/rejected_files.html', {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })

def download_rejected_files_excel(request):
    # Filter rejected files
    files = (
        Files.objects.filter(
            Q(office=request.user.office, processed=True, dept_approved=False)
        )
        .order_by('-dept_approved_at')
    )

    # Create a workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected Files"

    # Define headers
    headers = [
        "Filename",
        "Processed",
        "Remark",
        "Dept Approved",
        "Dept Rejected At"
    ]
    ws.append(headers)

    # Populate data rows
    for file in files:
        ws.append([
            file.filename,
            "Yes" if file.processed else "No",
            file.remark or "",
            "Yes" if file.dept_approved else "No",
            file.dept_approved_at.strftime('%Y-%m-%d %H:%M') if file.dept_approved_at else ""
        ])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{request.user.office.office_name}_rejected.xlsx"'
    wb.save(response)
    return response

@login_required
def rejected_by_digr(request):
    search_query = request.GET.get('q', '')

    files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved =False).order_by('uploaded_at')
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/rejected_by_digr.html', {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })


@login_required
def rejected_by_dist_a(request):
    search_query = request.GET.get('q', '')

    files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = False).order_by('uploaded_at')
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/rejected_by_dist_a.html', {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })

@login_required
def files_for_qc(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(office=request.user.office,processed=True,send_to_sro = True, dept_approved=None, district_rgtr_approved=None,digr_approved=None).order_by('-uploaded_at')
    if search_query:
        files = files.filter(filename__icontains=search_query) 
        
    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 

    return render(request, 'documents/files_for_qc.html', {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query, 
    })

@login_required
def approve_file(request, file_id):
    if request.user.is_department:
        
        file = Files.objects.get(id = file_id)
        previous = file.dept_approved
        print(previous)
        file.dept_approved = True
        file.dept_approved_by = request.user
        file.dept_approved_at = timezone.now()  
        file.save()
        
        UserActivity.objects.create(
                    user=request.user,
                    action=f"Approved document '{file.filename}'",
                    category='success',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )
        
        messages.success(request, f"File '{file.filename}' has been approved.")


        next_file = Files.objects.filter(office = request.user.office,processed=True,send_to_sro = True, dept_approved=None,district_rgtr_approved=None,digr_approved=None).order_by('-uploaded_at').first()


            
        if next_file:
            if previous == None:
                return redirect('processed_file', file_id=next_file.id)
            else:
                return redirect('rejected_files')
        else:
            return redirect('files_for_qc')
    else:
        messages.error(request, "You do not have permission to approve this file.")
    return redirect('files_for_qc')


@login_required
def approve_file_district_admin(request, file_id):
    if request.user.is_district_rgtr:
        file = Files.objects.get(id = file_id)
        file.dept_approved=True
        file.district_rgtr_approved = True
        file.district_rgtr_approved_by = request.user
        file.district_rgtr_approved_at = timezone.now()  
        file.save()
        
        UserActivity.objects.create(
                    user=request.user,
                    action=f"Approved document '{file.filename}'",
                    category='success',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )
        messages.success(request, f"File '{file.filename}' has been approved.")
        next_file = Files.objects.filter(office__district=request.user.district,processed=True, dept_approved=True,district_rgtr_approved=None,digr_approved=None).order_by('-dept_approved_at').first()
            
        if next_file:
            
            return redirect('sro_approved_file', file_id=next_file.id)  # Redirect to the next file
        else:
            return redirect('sro_approved_files')
    else:
        messages.error(request, "You do not have permission to approve this file.")
    return redirect('sro_approved_files')


@login_required
def approve_file_digr(request, file_id):
    if request.user.is_digr:
        file = Files.objects.get(id = file_id)
        file.dept_approved=True
        file.district_rgtr_approved = True
        file.digr_approved = True
        file.digr_approved_by = request.user
        file.digr_approved_at = timezone.now()  
        file.save()
        
        UserActivity.objects.create(
                    user=request.user,
                    action=f"Approved document '{file.filename}'",
                    category='danger',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )
        messages.success(request, f"File '{file.filename}' has been approved.")
        next_file = Files.objects.filter(office__district__zone =request.user.zone,processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=None).order_by('-district_rgtr_approved_at').first()
        
        if next_file:
            return redirect('d_a_approved_file', file_id=next_file.id)  # Redirect to the next file
        else:
            return redirect('d_a_approved_files')
    else:
        messages.error(request, "You do not have permission to approve this file.")
    return redirect('d_a_approved_files')

@login_required
def download_selected_files(request):
    if request.method == "POST":
        file_ids = request.POST.getlist("file_ids")
        files = Files.objects.filter(id__in=file_ids)

       
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for file in files:
                file_path = file.processed_file.path
                zip_file.write(file_path, file.processed_file.name) 

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="selected_files.zip"'
        return response
    else:
        return HttpResponse("Invalid request", status=400)
    

@login_required
def self_approved_files_sro(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(office = request.user.office,processed=True, dept_approved=True)
    
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Filter by date range
    if start_date:
        files = files.filter(dept_approved_at__date__gte=start_date)
    if end_date:
        files = files.filter(dept_approved_at__date__lte=end_date)
    
    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/self_approved_files_sro.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
        'start_date':start_date,
        'end_date':end_date,
    })

@login_required
def self_approved_files_sro_excel(request):
    search_query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    files = Files.objects.filter(
        office=request.user.office,
        processed=True,
        dept_approved=True,
    )

    if search_query:
        files = files.filter(filename__icontains=search_query)

    if start_date:
        files = files.filter(dept_approved_at__date__gte=start_date)
    if end_date:
        files = files.filter(dept_approved_at__date__lte=end_date)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Self Approved Files"

    headers = ['Filename','Year', 'Approved By', 'Approved On']
    ws.append(headers)

    for f in files:
        ws.append([
            f.filename,
            f.get_year() if hasattr(f, 'get_year') else '',
            f.dept_approved_by.get_full_name() if f.dept_approved_by else '',
            f.dept_approved_at.strftime('%Y-%m-%d') if f.dept_approved_at else '',
        ])

    # Write to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="self_approved_files.xlsx"'
    return response

@login_required
def all_dept_approved_files(request):
    search_query = request.GET.get('q', '')
    files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=True).order_by('uploaded_at')  
    if search_query:
        files = files.filter(filename__icontains=search_query) 

    paginator = Paginator(files, 10) 
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number) 
    return render(request, 'documents/all_dept_approved_files.html',  {
        'page_obj': page_obj,
        'files':files,
        'search_query': search_query,
    })

@login_required
def final_file(request, file_id):
    file = Files.objects.get(id = file_id)
    index_files = Files.objects.filter(processed = True, dept_approved = True,district_rgtr_approved = True,digr_approved=True).order_by('uploaded_at')  
    
    return render(request, 'documents/final_file.html', {'file': file,'index_files':index_files})


def overall_report(request):
    # Get the selected office from the request
    district_id = request.GET.get('district_id')  # Get the district_id from the query parameters
    
    # Create a base query to filter files
    files = Files.objects.annotate(
        upload_date=TruncDate('uploaded_at')  # Extract date part from uploaded_at
    ).values(
        'upload_date',
        'uploaded_by__office__district__zone__zone_name',
        'uploaded_by__office__district__district_name',
        'uploaded_by__office__office_name',
    ).annotate(
        submitted_count=Count('id'),
        dept_approved_count=Count('dept_approved', filter=Q(dept_approved=True)),
        dept_rejected_count=Count('dept_approved', filter=Q(dept_approved=False)),
    ).order_by('upload_date', 'uploaded_by__office__district__zone__zone_name')

    # If district_id is provided, filter the files by the selected office
    if district_id:
        files = files.filter(uploaded_by__office=district_id)

    # Paginate the files
    paginator = Paginator(files, 10)  # Show 10 files per page
    page_number = request.GET.get('page')  # Get the current page number from the query string
    page_obj = paginator.get_page(page_number)

    # Prepare chart data for each date's counts
    chart_data = {
        "dates": [],
        "submitted": [],
        "dept_approved": [],
        "dept_rejected": [],
    }

    for file in page_obj:
        chart_data["dates"].append(file['upload_date'].strftime('%Y-%m-%d'))  # Add formatted date
        chart_data["submitted"].append(file['submitted_count'])        
        chart_data["dept_approved"].append(file['dept_approved_count'])
        chart_data["dept_rejected"].append(file['dept_rejected_count'])

    # Return data to the template
    offices = Office.objects.all()
    context = {
        'files': page_obj,
        'chart_data': chart_data,
        'offices': offices,
    }
    return render(request, 'documents/overall_report.html', context)



def igr_report(request):
    # Get filter parameters from the request
    sro_name = request.GET.get('sro_name', '')
    district_name = request.GET.get('district_name', '')
    zone_name = request.GET.get('zone_name', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Apply filters
    filters = Q(send_to_sro=True)

    if sro_name:
        filters &= Q(office__office_name__icontains=sro_name)

    if district_name:
        filters &= Q(office__district__district_name__icontains=district_name)

    if zone_name:
        filters &= Q(office__district__zone__zone_name__icontains=zone_name)

    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.strptime(end_date, "%Y-%m-%d")
            filters &= Q(uploaded_at__date__range=(start_date, end_date))
        except ValueError:
            pass

    # Group data at different levels

    # 1. SRO Level Report
    sro_report = Files.objects.filter(filters).values(
        'office__office_name'
    ).annotate(
        total_qc=Count('id'),
        approved_count=Count('id', filter=Q(dept_approved=True)),
        rejected_count=Count('id', filter=Q(Q(dept_approved=False))),
        pending_count = Count('id', filter=Q(Q(dept_approved=None))),
        send_to_qc_pages=Sum('page_count'),

        approved_pages=Sum('page_count', filter=Q(dept_approved=True)),
        rejected_pages=Sum('page_count', filter=Q(dept_approved=False)),
        pending_pages=Sum('page_count', filter=Q(dept_approved=None)),
    ).order_by('office__office_name')

    sro_totals = Files.objects.filter(filters).aggregate(
                        total_qc=Count('id'),
                        approved_count=Count('id', filter=Q(dept_approved=True)),
                        rejected_count=Count('id', filter=Q(dept_approved=False)),
                        pending_count=Count('id', filter=Q(dept_approved=None)),
                        send_to_qc_pages=Sum('page_count'),
                        approved_pages=Sum('page_count', filter=Q(dept_approved=True)),
                        rejected_pages=Sum('page_count', filter=Q(dept_approved=False)),
                        pending_pages=Sum('page_count', filter=Q(dept_approved=None)),
                    )


    sro_data = [
        {
            "s_no": index,
            "sro_name": item['office__office_name'],
            "total_qc": item['total_qc'],
            "approved": item['approved_count'],
            "rejected": item['rejected_count'],
            "pending": item['pending_count'],
            "approved_pages": item.get("approved_pages") or 0,
            "rejected_pages": item.get("rejected_pages") or 0,
            "pending_pages": item.get("pending_pages") or 0,
            "send_to_qc_pages": item.get("send_to_qc_pages") or 0,
        }
        for index, item in enumerate(sro_report, start=1)
    ]

    # 2. District Level Report
    district_report = Files.objects.filter(filters).values(
        'office__district__district_name'
    ).annotate(
        total_qc=Count('id'),
        approved_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),
        rejected_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=False)),
        pending_count = Count('id', filter=Q(dept_approved=True, district_rgtr_approved=None)),
        total_count=Count('id', filter=Q(dept_approved=True)),
        send_to_qc_pages=Sum('page_count', filter=Q(dept_approved=True)),

        approved_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True)),
        rejected_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=False)),
        pending_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=None)),

    ).order_by('office__district__district_name')
    
    district_totals = Files.objects.filter(filters).aggregate(
                        total_qc=Count('id'),
                        total_count=Count('id', filter=Q(dept_approved=True)),

                        approved_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),
                        rejected_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=False)),
                        pending_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=None)),

                        send_to_qc_pages=Sum('page_count', filter=Q(dept_approved=True)),
                        approved_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True)),
                        rejected_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=False)),
                        pending_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=None)),
                    )

    district_data = [
        {
            "s_no": index,
            "district": item['office__district__district_name'],
            "total_qc": item['total_qc'],
            "approved": item['approved_count'],
            "rejected": item['rejected_count'],
            "pending": item['pending_count'],
            "total_count":item["total_count"],
            "approved_pages": item.get("approved_pages") or 0,
            "rejected_pages": item.get("rejected_pages") or 0,
            "pending_pages": item.get("pending_pages") or 0,
            "send_to_qc_pages": item.get("send_to_qc_pages") or 0,
        }
        for index, item in enumerate(
        [d for d in district_report if d["total_count"] > 0], start=1
    )
    ]

    # 3. Zone Level Report
    zone_report = Files.objects.filter(filters).values(
                        'office__district__zone__zone_name'
                    ).annotate(
                        total_qc=Count('id'),
                        approved_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                        rejected_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                        pending_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),
                        total_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),
                        send_to_qc_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True)),
                        # Page counts
                        approved_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                        rejected_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                        pending_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None))
                    ).order_by('office__district__zone__zone_name')
    
    zone_totals = Files.objects.filter(filters).aggregate(
                    total_qc=Count('id'),

                    total_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),

                    approved_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                    rejected_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                    pending_count=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),

                    send_to_qc_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True)),

                    approved_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                    rejected_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                    pending_pages=Sum('page_count', filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),
                )

    zone_data = [
        {
            "s_no": index,
            "zone": item['office__district__zone__zone_name'],
            "total_qc": item['total_qc'],
            "approved": item['approved_count'],
            "rejected": item['rejected_count'],
            "pending": item['pending_count'],
            "total_count":item["total_count"],
            "approved_pages": item.get("approved_pages") or 0,
            "rejected_pages": item.get("rejected_pages") or 0,
            "pending_pages": item.get("pending_pages") or 0,
            "send_to_qc_pages": item.get("send_to_qc_pages") or 0,
        }
        for index, item in enumerate(
        [z for z in zone_report if z["total_count"] > 0], start=1
    )
    ]
    # filtered_zone_data = [row for row in zone_data if row.total_count > 0]
    # Get lists for dropdown filters
    sro_list = Files.objects.values_list("office__office_name", flat=True).distinct()
    district_list = Files.objects.values_list("office__district__district_name", flat=True).distinct()
    zone_list = Files.objects.values_list("office__district__zone__zone_name", flat=True).distinct()

    return render(request, 'documents/igr_report.html', {
        'sro_data': sro_data,
        'sro_totals' : sro_totals ,
        'district_data': district_data,
        'district_totals':district_totals,
        'zone_data': zone_data,
        'zone_totals':zone_totals,
        'sro_list': sro_list,
        'district_list': district_list,
        'zone_list': zone_list,
        'sro_name': sro_name,
        'district_name': district_name,
        'zone_name': zone_name,
        'start_date': start_date,
        'end_date': end_date
    })

def download_sro_report(request):
    sro_name = request.GET.get("sro_name", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    # Apply filters
    filters = Q(send_to_sro=True)
    if sro_name:
        filters &= Q(office__office_name=sro_name)
    if start_date:
        filters &= Q(uploaded_at__gte=start_date)
    if end_date:
        filters &= Q(uploaded_at__lte=end_date)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SRO Report"

    # Headers
    headers = [
         "SRO Name", "Total Docs Sent to QC","Total Pages Sent to QC", "Approved Document","Approved Document Pages",
          "Rejected Document","Rejected Document Pages", "Pending Document","Pending Document Pages"
          
    ]
    ws.append(headers)


    # Fetch Data
    sro_report = Files.objects.filter(filters).values("office__office_name").annotate(
                        total_qc=Count("id"),
                        approved_count=Count("id", filter=Q(dept_approved=True)),
                        rejected_count=Count("id", filter=Q(dept_approved=False)),
                        pending_count=Count("id", filter=Q(dept_approved=None)),
                        approved_pages=Sum("page_count", filter=Q(dept_approved=True)),
                        rejected_pages=Sum("page_count", filter=Q(dept_approved=False)),
                        pending_pages=Sum("page_count", filter=Q(dept_approved=None)),
                        send_to_qc_pages=Sum("page_count")
                    )


    for index, office in enumerate(sro_report, start=1):
        ws.append([
            office["office__office_name"] or "",
            office["total_qc"],
            office["send_to_qc_pages"] or 0,
            office["approved_count"],
            office["approved_pages"] or 0,
            office["rejected_count"],
            office["rejected_pages"] or 0,
            office["pending_count"],
            office["pending_pages"] or 0,
            
        ])


    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="SRO_Report.xlsx"'

    wb.save(response)
    return response


def download_district_report(request):
    district_name = request.GET.get("district_name", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    # Apply filters
    filters = Q(send_to_sro=True)
    if district_name:
        filters &= Q(office__district__district_name=district_name)
    if start_date:
        filters &= Q(uploaded_at__gte=start_date)
    if end_date:
        filters &= Q(uploaded_at__lte=end_date)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "District Report"

    # Headers
    headers = [
             "District", "Total Documents Sent to QC","Total Document Pages Sent to QC", "Approved Document","Approved Document Pages",
          "Rejected Document","Rejected Document Pages", "Pending Document","Pending Document Pages"
                ]
    ws.append(headers)


    # Fetch Data
    district_report = Files.objects.filter(filters).values("office__district__district_name").annotate(
                total_qc=Count("id", filter=Q(dept_approved=True)),
                approved_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=True)),
                rejected_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=False)),
                pending_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=None)),
                approved_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=True)),
                rejected_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=False)),
                pending_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=None)),
                send_to_qc_pages=Sum("page_count", filter=Q(dept_approved=True))
            )


    for index, district in enumerate(district_report, start=1):
        if district["total_qc"] > 0:
            ws.append([
                district["office__district__district_name"] or "",
                district["total_qc"],
                district["send_to_qc_pages"] or 0,
                district["approved_count"],
                district["approved_pages"] or 0,
                district["rejected_count"],
                district["rejected_pages"] or 0,
                district["pending_count"],
                district["pending_pages"] or 0
            ])


    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="District_Report.xlsx"'

    wb.save(response)
    return response

def download_zone_report(request):
    zone_name = request.GET.get("zone_name", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    # Apply filters
    filters = Q(send_to_sro=True)
    if zone_name:
        filters &= Q(office__district__zone__zone_name=zone_name)
    if start_date:
        filters &= Q(uploaded_at__gte=start_date)
    if end_date:
        filters &= Q(uploaded_at__lte=end_date)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zone Report"

    # Headers
    headers = [
         "Zone", "Total Document Sent to QC","Total Document Pages Sent to QC", "Approved Document","Approved Document Pages",
          "Rejected Document","Rejected Document Pages", "Pending Document","Pending Document Pages"
    ]
    ws.append(headers)


    # Fetch Data
    zone_report = Files.objects.filter(filters).values("office__district__zone__zone_name").annotate(
                    total_qc=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=True)),
                    approved_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                    rejected_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                    pending_count=Count("id", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),
                    approved_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=True)),
                    rejected_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=False)),
                    pending_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=True, digr_approved=None)),
                    send_to_qc_pages=Sum("page_count", filter=Q(dept_approved=True, district_rgtr_approved=True))
                )


    for index, zone in enumerate(zone_report, start=1):
        if zone["total_qc"] > 0:
            ws.append([
                zone["office__district__zone__zone_name"] or "",
                zone["total_qc"],
                zone["send_to_qc_pages"] or 0,
                zone["approved_count"],
                zone["approved_pages"] or 0,
                zone["rejected_count"],
                zone["rejected_pages"] or 0,
                zone["pending_count"],
                zone["pending_pages"] or 0
            ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Zone_Report.xlsx"'

    wb.save(response)
    return response


def download_excel_report(request):
    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Files Report"

    # Define column headers with Office Name added
    headers = [
        "Office Name", "Filename","Page Count", "Uploaded By", "Uploaded At",
        "Processed", "Remark", "Remark by Scanner", "Send to SRO",
        "Dept Approved", "Dept Approved By", "Dept Approved At",
        "District RGTR Approved", "District RGTR Approved By", "District RGTR Approved At",
        "DIGR Approved", "DIGR Approved By", "DIGR Approved At"
    ]
    ws.append(headers)

    # Query all Files and populate rows
    files = Files.objects.all()
    for file in files:
        row = [
            file.office.office_name if file.office else "", 
            file.filename,
            file.page_count,
            file.uploaded_by.username if file.uploaded_by else "",
            timezone.localtime(file.uploaded_at).strftime('%Y-%m-%d %H:%M:%S') if file.uploaded_at else "",
            file.processed,
            file.remark,
            file.remark_by_scanner,
            file.send_to_sro,
            file.dept_approved,
            file.dept_approved_by.username if file.dept_approved_by else "",
            file.dept_approved_at.strftime('%Y-%m-%d %H:%M:%S') if file.dept_approved_at else "",
            file.district_rgtr_approved,
            file.district_rgtr_approved_by.username if file.district_rgtr_approved_by else "",
            file.district_rgtr_approved_at.strftime('%Y-%m-%d %H:%M:%S') if file.district_rgtr_approved_at else "",
            file.digr_approved,
            file.digr_approved_by.username if file.digr_approved_by else "",
            file.digr_approved_at.strftime('%Y-%m-%d %H:%M:%S') if file.digr_approved_at else "",
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Create an HTTP response with the Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="files_report.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response


@login_required
def add_received_docs(request):
    if request.method == 'POST':
        # Get data from the form
        date = request.POST.get('from_date')
        document_type = request.POST.getlist('document_type')
        index_type = request.POST.getlist('index_type')
        received = request.POST.get('received')
        remark = request.POST.get('remark')

        # Create a new ReceivedDocuments object
        document = ReceivedDocuments(
            date=date,
            document_type=','.join(document_type),  # Join selected document types into a comma-separated string
            index_type=','.join(index_type),  # Join selected index types into a comma-separated string
            received=received,
            remark=remark,
            created_by=request.user
        )
        document.save()
        return redirect('add_received_docs') 
    
    documents = ReceivedDocuments.objects.filter(created_by=request.user)
    context = {
        'documents': documents
    }
    return render(request, 'documents/received_documents.html', context)


@login_required
def admin_approve_file(request, file_id):
    if request.user.is_admin  or  request.user.is_agency_qc_employee or request.user.is_agency_admin:
        
        file = Files.objects.get(id = file_id)
        previous = file.admin_approved
        file.admin_approved = True
        file.admin_approved_by = request.user
        file.admin_approved_at = timezone.now()
        file.send_to_sro = True  
        file.save()
        
        UserActivity.objects.create(
                    user=request.user,
                    action=f"Approved document '{file.filename}'",
                    category='success',
                    ip_address=request.ip_address,  
                    device_type=request.device_type,  
                    browser=request.browser,  
                    os=request.os 
                )
        
        messages.success(request, f"File '{file.filename}' has been approved.")

        next_file = Files.objects.filter(processed=True,send_to_sro = False,send_to_qc = True, dept_approved=None,district_rgtr_approved=None,digr_approved=None).order_by('-uploaded_at').first()
        if request.user.is_agency_qc_employee or request.user.is_agency_admin:
            next_file = Files.objects.filter(uploaded_by__agency = request.user.agency ,processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at').first()
            
        if next_file:
            if previous == None:
                return redirect('file', file_id=next_file.id)
            else:
                return redirect('processed_files')
        else:
            return redirect('processed_files')
    else:
        messages.error(request, "You do not have permission to approve this file.")
    return redirect('processed_files')


@login_required
def file(request, file_id):
    file = get_object_or_404(Files, id=file_id)

    if request.method == 'POST':
        if request.user.is_admin  or  request.user.is_agency_qc_employee or request.user.is_agency_admin:
            remark = request.POST.get('remark', '')

            # Append admin username to the remark
            remark_with_username = f"({request.user.username}) {remark}"
            file.remark = remark_with_username
            file.admin_approved = False
            file.admin_approved_by = request.user
            file.admin_approved_at = timezone.now()
            file.save()

            # Log user activity with IP, device, browser, and OS
            UserActivity.objects.create(
                user=request.user,
                action=f"Rejected document '{file.filename}'",
                category='danger',
                ip_address=request.ip_address,  
                device_type=request.device_type,  
                browser=request.browser,  
                os=request.os  
            )

            messages.success(request, f"{file.filename} has been Rejected.")

            # Fetch next file for processing
            next_file = Files.objects.filter(
                processed=True, send_to_sro=False, send_to_qc=True,
                admin_approved=None, dept_approved=None,
                district_rgtr_approved=None, digr_approved=None
            ).order_by('-uploaded_at').first()
            
            if request.user.is_agency_qc_employee or request.user.is_agency_admin:
                next_file = Files.objects.filter(uploaded_by = request.user,processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at').first()

            if next_file:
                return redirect('file', file_id=next_file.id)
            return redirect('processed_files')

        else:
            messages.error(request, "You are not authorized to perform this action.")
            return redirect('processed_files')

    # Group files by uploader for display
    grouped_files = Files.objects.filter(
        processed=True, send_to_qc=True, send_to_sro=False, dept_approved=None
    ).order_by('-uploaded_at')
    
    if request.user.is_agency_qc_employee or request.user.is_agency_admin:
        grouped_files = Files.objects.filter(uploaded_by = request.user ,processed=True,send_to_sro=False,send_to_qc = True,admin_approved = None,dept_approved=None).order_by('-uploaded_at')

    index_files = {
        user: list(files) for user, files in groupby(grouped_files, key=attrgetter('uploaded_by'))
    }

    return render(request, 'documents/file.html', {'file': file, 'index_files': index_files})

@login_required
def zone_wise_report(request):
    # 1) Get optional zone filter
    zone_filter = request.GET.get('zone', '').strip()

    # 2) Base queryset of Files → group by zone
    qs = Files.objects.filter(office__isnull=False)
    if zone_filter:
        qs = qs.filter(office__district__zone__zone_name__icontains=zone_filter)

    zone_stats = (
        qs.values(zone_name=F('office__district__zone__zone_name'))
          .annotate(
              # SRO stage
              sro_given     = Count('id', filter=Q(processed=True)),
              sro_approved  = Count('id', filter=Q(processed=True, dept_approved=True)),
              # DR stage
              dr_given      = Count('id', filter=Q(dept_approved=True)),
              dr_approved   = Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),
              # DIGR stage
              digr_given    = Count('id', filter=Q(district_rgtr_approved=True)),
              digr_approved = Count('id', filter=Q(district_rgtr_approved=True, digr_approved=True)),
          )
          .order_by('zone_name')
    )
    
    zone_totals = qs.aggregate(
                                sro_given=Count('id', filter=Q(processed=True)),
                                sro_approved=Count('id', filter=Q(processed=True, dept_approved=True)),

                                dr_given=Count('id', filter=Q(dept_approved=True)),
                                dr_approved=Count('id', filter=Q(dept_approved=True, district_rgtr_approved=True)),

                                digr_given=Count('id', filter=Q(district_rgtr_approved=True)),
                                digr_approved=Count('id', filter=Q(district_rgtr_approved=True, digr_approved=True)),
                            )

    # 3) If ?export=excel, return an Excel file
    if request.GET.get('export') == 'excel':
        return _export_zone_stats_to_excel(zone_stats, zone_filter)

    # 4) Otherwise render HTML
    return render(request, 'documents/zonewise_report.html', {
        'zone_stats': zone_stats,
        'zone_totals' : zone_totals,
        'zone_filter': zone_filter,
    })


def _export_zone_stats_to_excel(zone_stats, zone_filter):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Zone Report'

    # Header row
    headers = [
        'Zone',
        'QC Given by SRO', 'QC Approved by SRO',
        'QC Given by DR',  'QC Approved by DR',
        'QC Given by DIGR','QC Approved by DIGR',
    ]
    ws.append(headers)

    # Data rows
    for z in zone_stats:
        ws.append([
            z['zone_name'],
            z['sro_given'],   z['sro_approved'],
            z['dr_given'],    z['dr_approved'],
            z['digr_given'],  z['digr_approved'],
        ])

    # Optional: auto‑size columns
    for col in ws.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = length + 2

    # Build response
    filename = 'zone_report'
    if zone_filter:
        filename += f'_{zone_filter}'
    filename += '.xlsx'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


class UploadTiffFileView(CreateAPIView):
    queryset = Files.objects.all()
    serializer_class = FilesUploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
@login_required
def agency_wise_users_data(request, user_id):
    user = User.objects.get(id=user_id)
    files_qs = Files.objects.filter(uploaded_by=user).order_by('-uploaded_at')

    # Filter by approval status
    admin_approved = request.GET.get('admin_approved')
    dept_approved = request.GET.get('dept_approved')
    dr_approved = request.GET.get('dr_approved')
    digr_approved = request.GET.get('digr_approved')

    if admin_approved in ['true', 'false']:
        files_qs = files_qs.filter(admin_approved=(admin_approved == 'true'))
    if dept_approved in ['true', 'false']:
        files_qs = files_qs.filter(dept_approved=(dept_approved == 'true'))
    if dr_approved in ['true', 'false']:
        files_qs = files_qs.filter(district_rgtr_approved=(dr_approved == 'true'))
    if digr_approved in ['true', 'false']:
        files_qs = files_qs.filter(digr_approved=(digr_approved == 'true'))

    # ✅ Filter by date span
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        files_qs = files_qs.filter(uploaded_at__date__gte=parse_date(start_date))
    if end_date:
        files_qs = files_qs.filter(uploaded_at__date__lte=parse_date(end_date))


    # Excel download
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Filename", "Page Count", "Admin Approved", "Dept Approved", "DR Approved", "DIGR Approved", "Remark"])
        for f in files_qs:
            ws.append([
                f.filename, f.page_count,
                f.admin_approved, f.dept_approved,
                f.district_rgtr_approved, f.digr_approved,
                f.remark
            ])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="user_{user_id}_files.xlsx"'
        wb.save(response)
        return response

    total_pages = files_qs.aggregate(total_pages=Sum('page_count'))['total_pages'] or 0
    
    approved_admin = files_qs.filter(admin_approved=True).count()
    rejected_admin = files_qs.filter(admin_approved=False).count()

    approved_dept = files_qs.filter(dept_approved=True).count()
    rejected_dept = files_qs.filter(dept_approved=False).count()

    approved_dr = files_qs.filter(district_rgtr_approved=True).count()
    rejected_dr = files_qs.filter(district_rgtr_approved=False).count()

    approved_digr = files_qs.filter(digr_approved=True).count()
    rejected_digr = files_qs.filter(digr_approved=False).count()


    context = {
        'files': files_qs,
        'user': user,
        'total': files_qs.count(),
        'start_date': start_date,
        'end_date': end_date,
        'summary': {
                    'total_pages': total_pages,
                    'approved_admin': approved_admin,
                    'rejected_admin': rejected_admin,
                    'approved_dept': approved_dept,
                    'rejected_dept': rejected_dept,
                    'approved_dr': approved_dr,
                    'rejected_dr': rejected_dr,
                    'approved_digr': approved_digr,
                    'rejected_digr': rejected_digr,
                    }
    }
    return render(request, "documents/agency_wise_users_data.html", context)


@login_required
@require_POST
def zip_download_view(request):
    file_ids = request.POST.getlist('file_ids')
    files = Files.objects.filter(id__in=file_ids, processed_file__isnull=False)
    
    file = files.first()
    name = 'No Files selected'
    if file:
        name  = file.uploaded_by.username
    
    tmp = tempfile.TemporaryFile()
    with zipfile.ZipFile(tmp, 'w') as zipf:
        for f in files:
            if os.path.exists(f.processed_file.path):
                zipf.write(f.processed_file.path, arcname=os.path.basename(f.processed_file.name))
    tmp.seek(0)

    response = FileResponse(tmp, as_attachment=True, filename=f'{name}.zip')
    return response

from django.http import JsonResponse
import os
import re
from collections import defaultdict
from django.db import IntegrityError, transaction
from django.db.models import Min

def remove(request):
    pattern_b_regex = re.compile(r"^(R_\d+_\d+_\d+_\d{4})")
    deleted = []

    # Precompute all required data in optimized queries
    file_data = Files.objects.values_list('id', 'filename', 'processed_file', 'uploaded_at')
    existing_filenames = set(Files.objects.values_list('filename', flat=True))

    # Group files by base filename in memory
    file_groups = defaultdict(list)
    path_to_file = {}
    for file_id, filename, processed_file, uploaded_at in file_data:
        file_name = os.path.basename(processed_file)
        base_filename = os.path.splitext(file_name)[0]
        
        if not pattern_b_regex.match(base_filename):
            continue
            
        file_groups[base_filename].append((file_id, filename, processed_file, uploaded_at))
        path_to_file[file_id] = (filename, processed_file)

    # Process groups in bulk
    to_delete_ids = set()
    to_update = {}
    
    for base_filename, group in file_groups.items():
        group_sorted = sorted(group, key=lambda x: x[3] or x[0])  # uploaded_at or id
        
        # Handle filename conflict
        if base_filename in existing_filenames:
            for file_id, filename, processed_file, _ in group_sorted:
                if filename != base_filename:
                    to_delete_ids.add(file_id)
            continue

        # Select keeper (oldest file)
        keeper_id, keeper_filename, keeper_processed, _ = group_sorted[0]
        if keeper_filename != base_filename:
            to_update[keeper_id] = base_filename
            existing_filenames.add(base_filename)  # Update cache

        # Mark duplicates for deletion
        for file_id, _, processed_file, _ in group_sorted[1:]:
            to_delete_ids.add(file_id)

    # Execute bulk operations
    with transaction.atomic():
        # Bulk update filenames
        if to_update:
            update_objs = []
            for file_id, new_name in to_update.items():
                update_objs.append(Files(id=file_id, filename=new_name))
            Files.objects.bulk_update(update_objs, ['filename'])
        
        # Bulk delete database records
        if to_delete_ids:
            delete_files = Files.objects.filter(id__in=to_delete_ids)
            deleted_paths = [f.processed_file.name for f in delete_files]
            delete_files.delete()
            deleted.extend(deleted_paths)

    # Handle file deletions (outside transaction)
    for file_id in to_delete_ids:
        try:
            _, processed_file = path_to_file[file_id]
            if processed_file:
                abs_path = processed_file.path
                if os.path.exists(abs_path):
                    os.remove(abs_path)
        except Exception:
            pass

    return JsonResponse({
        "status": "success",
        "deleted_duplicates": deleted,
        "total_deleted": len(deleted),
    })


import io
import os
import re
import zipfile

from django.http import StreamingHttpResponse
from django.views import View
from django.contrib.auth.mixins import UserPassesTestMixin

from .models import Files


class AllOfficesZipExportView(UserPassesTestMixin, View):
    """
    GET /export-zip-all/
    Returns ONE ZIP that contains every processed file, arranged as:
        <office_code>/<file_type>/<year>/<original_filename>
    """

    chunk_size = 8192  # streaming chunk size

    # ── permission ─────────────────────────────────────────────────────────
    def test_func(self):
        return self.request.user.is_staff          # staff-only endpoint

    # ── GET handler ────────────────────────────────────────────────────────
    def get(self, request):
        qs = (
            Files.objects
                 .select_related("office")
                 .filter(office__isnull=False,
                         processed_file__isnull=False)
                 .only("filename", "processed_file", "office__office_code")
        )

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:

            # Use iterator() → keeps memory footprint small even with 90k+ rows
            for f in qs.iterator(chunk_size=2000):
                if not f.filename:
                    continue

                # 1) cleanse the raw filename
                fname = os.path.basename(
                    f.filename.replace("_processed", "").replace("processed", "")
                )

                # 2) physical path on disk; skip if missing
                disk_path = f.processed_file.path
                if not os.path.exists(disk_path):
                    continue

                # 3) build the relative path inside the ZIP
                office_code = f.office.office_code or "UNKNOWN"
                inner_path = (
                    f"{office_code}/"                 # top-level office folder
                    f"{self._structured_path(fname)}" #   type/year/filename
                )

                # 4) add the file
                zf.write(disk_path, inner_path)

        # ── stream the ZIP to client ───────────────────────────────────────
        zip_buffer.seek(0)
        response = StreamingHttpResponse(
            streaming_content=iter(
                lambda: zip_buffer.read(self.chunk_size), b""
            ),
            content_type="application/zip",
        )
        response["Content-Disposition"] = 'attachment; filename="all_offices.zip"'
        return response

    # ── helper: compute "<file_type>/<year>/<filename>" ────────────────────
    @staticmethod
    def _structured_path(filename: str) -> str:
        """
        For a cleaned filename, return   <file_type>/<year>/<filename>
        Falls back to UNKNOWN/UNKNOWN/filename if pattern not recognised.
        """
        patterns = [
            # Index: I_<Office>_<IdxType>_<Year>_<Vol>_P?
            (r"^(I)_[\d_]+_(\d{4})_",                        1, 2),
            # MTPR: MTPR_<Office>_<Vol>_P?
            (r"^(MTPR)_[\d_]+_P?\d+$",                       1, None),
            # RH: RH_<Office>_<Vol>_P?
            (r"^(RH)_[\d_]+_P?\d+$",                         1, None),
            # Docs: R|LO|MO|CO_<Office>_<Book>_<Run>_<Year>
            (r"^((?:R|LO|MO|CO))_[\d_]+_(\d{4})$",           1, 2),
        ]

        for pattern, type_grp, year_grp in patterns:
            m = re.match(pattern, filename)
            if m:
                doc_type = m.group(type_grp)
                year     = m.group(year_grp) if year_grp and m.group(year_grp) else "UNKNOWN"
                return f"{doc_type}/{year}/{filename}"

        # fallback
        return f"UNKNOWN/UNKNOWN/{filename}"
    


class FilePathDocumentDownloadURL(generics.GenericAPIView):
    permission_classes      = [AllowAny]
    authentication_classes  = (TokenAuthentication,)

    def get(self, request, *args, **kwargs):
        file_id = request.query_params.get("id")  # ?id=123

        try:
            file_obj = Files.objects.get(id=file_id)

            status_map = {True: "TRUE", False: "FALSE", None: "NULL"}
            response_data = {
                "file_path":      request.build_absolute_uri(file_obj.processed_file.url),
                "dept_approved":  status_map[file_obj.dept_approved],
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Files.DoesNotExist:
            raise Http404("Document with this ID does not exist")
        except Exception as exc:
            return Response({"message": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        
        
from django.db.models import Case, When, Value, BooleanField
   
def export_dept_approval_excel(request):
    """
    Export an Excel sheet of every *processed* file with its DB-ID,
    filename, and department-approval status (True / False / Null).
    """

    # ── 1  Query + normalise dept_approved to real True / False / None ──
    qs = (
        Files.objects
        .filter(processed=True)
        .annotate(
            approval_clean=Case(
                When(dept_approved=True,  then=Value(True)),
                When(dept_approved=False, then=Value(False)),
                default=Value(None),
                output_field=BooleanField(),
            )
        )
        .values_list("id", "filename", "approval_clean")   # ← include id
    )

    # ── 2  Build workbook in memory ─────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Dept Approval"
    ws.append(["ID", "Filename", "Dept Approved"])        # header

    display = {True: "True", False: "False", None: "Null"}

    for file_id, filename, approval in qs.iterator():
        ws.append([file_id, filename, display[approval]])

    # ── 3  Return as attachment ─────────────────────────────────────────
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dept_approval.xlsx"'
    return response


from django.utils.text import slugify
def stream_zip_file(queryset, zip_name="approved_files.zip"):
    """Generator to stream ZIP file content without storing entire ZIP in memory"""
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file_obj in queryset.iterator():  # Use `.iterator()` for memory efficiency
            if file_obj.processed_file and os.path.exists(file_obj.processed_file.path):
                # Clean filename and add to zip
                filename = slugify(os.path.basename(file_obj.processed_file.name))
                zip_file.write(file_obj.processed_file.path, arcname=filename)

    buffer.seek(0)
    response = StreamingHttpResponse(
        streaming_content=buffer,
        content_type='application/zip'
    )
    response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return response

def download_dept_approved_zip(request):
    """Django view to download a ZIP of dept-approved files"""
    # Filter dept_approved=True files
    approved_files = Files.objects.filter(dept_approved=True).only('processed_file')

    return stream_zip_file(approved_files)